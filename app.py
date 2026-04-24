"""
Dashboard de Conciliação MOTZ - Streamlit (v3)
Upload de PDFs Repom + MOTZ (XLSX) + ATUA (XLS) → conciliação → visualização
v3: cores da skill + distribuição clicável (cards + gráfico de pizza)
"""
import streamlit as st
import pandas as pd
import subprocess
import tempfile
import os
import shutil
from pathlib import Path
from datetime import datetime, timedelta
import hashlib
import io
import sys
import plotly.express as px

# ============================================================
# Logos (base64) — embutidos pra não depender de arquivos externos
# ============================================================
try:
    from logos_b64 import LOGO_REPOM, LOGO_MOTZ, LOGO_ATUA, LOGO_PIANETTO
except ImportError:
    try:
        from logos_b64 import LOGO_REPOM, LOGO_MOTZ, LOGO_ATUA
        LOGO_PIANETTO = ""
    except ImportError:
        # Fallback: sem logos (usa só o texto)
        LOGO_REPOM = LOGO_MOTZ = LOGO_ATUA = LOGO_PIANETTO = ""

# ============================================================
# Configuração da página
# ============================================================
st.set_page_config(
    page_title="Conciliação MOTZ",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ============================================================
# Paleta de cores — alinhada com a skill /conciliacao-motz
# ============================================================
# Verde  = OK (diferença até R$100 favorável à MOTZ)
# Vermelho = ATUA MAIOR > R$100 (crítico)
# Azul   = ATUA MAIOR até R$100 (diferença pequena)
# Amarelo = NÃO ENCONTRADO / sem PDF
# Roxo   = Saldo Aberto
COLORS = {
    # Paleta ajustada para fundo preto — fundos escuros saturados + textos claros
    "OK":            {"bg": "#1F4D1A", "fg": "#C3E8A8", "border": "#5AA84A", "plot": "#6BC54C"},
    "ATUA MAIOR":    {"bg": "#5A1A1A", "fg": "#F5B8B8", "border": "#E05757", "plot": "#E85858"},
    "ATUA MENOR":    {"bg": "#1A3A5A", "fg": "#B8D8F5", "border": "#5F95D0", "plot": "#5FA0E0"},
    "NAO ENCONTRADO":{"bg": "#5A3F0F", "fg": "#F5DB9E", "border": "#C08D3A", "plot": "#E0A040"},
    "SALDO ABERTO":  {"bg": "#3A2F6B", "fg": "#CFC4F5", "border": "#7A6FD0", "plot": "#8B7FE0"},
}

# CSS para aproximar do visual do dashboard HTML
st.markdown(f"""
<style>
    /* === TEMA ESCURO FORÇADO === */
    .stApp, [data-testid="stAppViewContainer"], [data-testid="stMain"], .main {{
        background-color: #0E0E0E !important;
        color: #EAEAEA !important;
    }}
    [data-testid="stHeader"] {{ background-color: #0E0E0E !important; }}
    [data-testid="stSidebar"] {{ background-color: #141414 !important; }}

    .main .block-container {{ padding-top: 2rem; padding-bottom: 3rem; max-width: 1400px; }}

    /* Tipografia */
    h1, h2, h3, h4, h5, h6, p, span, label {{ color: #EAEAEA; }}
    h1 {{ font-size: 22px !important; font-weight: 500 !important; margin-bottom: 0 !important; color: #F5F5F5 !important; }}
    h2 {{ font-size: 16px !important; font-weight: 500 !important; color: #F0F0F0 !important; }}
    h3 {{ font-size: 14px !important; font-weight: 500 !important; color: #F0F0F0 !important; }}
    small, [data-testid="stCaptionContainer"] {{ color: #9A9A9A !important; }}

    /* Containers com borda */
    [data-testid="stVerticalBlockBorderWrapper"] > div,
    div[data-testid="stExpander"] {{
        background-color: #1A1A1A !important;
        border-color: #2A2A2A !important;
    }}

    /* Métricas */
    .stMetric, [data-testid="stMetric"] {{
        background: #1A1A1A !important;
        border: 1px solid #2A2A2A;
        border-radius: 10px;
        padding: 12px 16px;
    }}
    .stMetric label, [data-testid="stMetricLabel"] {{ font-size: 12px !important; color: #9A9A9A !important; }}
    .stMetric [data-testid="stMetricValue"] {{ font-size: 22px !important; font-weight: 500 !important; color: #F5F5F5 !important; }}
    .stMetric [data-testid="stMetricDelta"] {{ color: #B0B0B0 !important; }}

    /* Inputs e seletores */
    input, textarea, select,
    [data-baseweb="input"] > div, [data-baseweb="select"] > div {{
        background-color: #1A1A1A !important;
        color: #EAEAEA !important;
        border-color: #333 !important;
    }}
    [data-baseweb="popover"], [data-baseweb="menu"] {{ background-color: #1A1A1A !important; }}

    /* Upload de arquivos */
    [data-testid="stFileUploader"] section {{
        background-color: #1A1A1A !important;
        border: 1px dashed #444 !important;
        border-radius: 10px;
        padding: 14px;
    }}

    /* Tabela (DataFrame) */
    .stDataFrame {{ font-size: 12px; background-color: #1A1A1A !important; }}
    .stDataFrame [data-testid="stTable"] {{ background-color: #1A1A1A !important; }}
    .stDataFrame table {{ color: #EAEAEA !important; }}
    .stDataFrame thead th {{
        background-color: #2A2A2A !important;
        color: #F5F5F5 !important;
        border-color: #3A3A3A !important;
    }}
    .stDataFrame tbody tr {{ background-color: #1A1A1A !important; }}
    .stDataFrame tbody td {{ border-color: #2A2A2A !important; }}

    /* Info/warning/success/error */
    [data-testid="stAlert"] {{
        background-color: #1A1A1A !important;
        color: #EAEAEA !important;
        border: 1px solid #2A2A2A !important;
    }}

    /* Botões */
    .stButton > button {{
        border-radius: 8px;
        font-size: 13px;
        padding: 6px 14px;
        background-color: #2A2A2A;
        color: #EAEAEA;
        border: 1px solid #3A3A3A;
    }}
    .stButton > button:hover {{ background-color: #3A3A3A; border-color: #4A4A4A; }}
    .stButton > button[kind="primary"] {{
        background-color: #378ADD !important;
        color: #FFFFFF !important;
        border-color: #378ADD !important;
    }}
    .stButton > button[kind="primary"]:hover {{ background-color: #2970BF !important; }}

    /* Download button */
    .stDownloadButton > button {{
        background-color: #2A2A2A !important;
        color: #EAEAEA !important;
        border: 1px solid #3A3A3A !important;
    }}

    /* Divider */
    hr {{ border-color: #2A2A2A !important; }}

    /* === BADGES DE STATUS === */
    .status-ok {{ background: {COLORS['OK']['bg']}; color: {COLORS['OK']['fg']}; padding: 2px 8px; border-radius: 999px; font-size: 11px; font-weight: 500; }}
    .status-maior {{ background: {COLORS['ATUA MAIOR']['bg']}; color: {COLORS['ATUA MAIOR']['fg']}; padding: 2px 8px; border-radius: 999px; font-size: 11px; font-weight: 500; }}
    .status-menor {{ background: {COLORS['ATUA MENOR']['bg']}; color: {COLORS['ATUA MENOR']['fg']}; padding: 2px 8px; border-radius: 999px; font-size: 11px; font-weight: 500; }}
    .status-ne {{ background: {COLORS['NAO ENCONTRADO']['bg']}; color: {COLORS['NAO ENCONTRADO']['fg']}; padding: 2px 8px; border-radius: 999px; font-size: 11px; font-weight: 500; }}
    .status-aberto {{ background: {COLORS['SALDO ABERTO']['bg']}; color: {COLORS['SALDO ABERTO']['fg']}; padding: 2px 8px; border-radius: 999px; font-size: 11px; font-weight: 500; }}

    /* === CARDS CLICÁVEIS DE DISTRIBUIÇÃO === */
    div[data-testid="stButton"] > button[kind="secondary"] {{
        width: 100%; border-radius: 10px; padding: 12px 14px; font-size: 13px; font-weight: 500;
        text-align: left; border: 1.5px solid transparent; transition: all 0.15s ease; min-height: 62px;
    }}
    div[data-testid="stButton"] > button[kind="secondary"]:hover {{ transform: translateY(-1px); box-shadow: 0 2px 8px rgba(0,0,0,0.35); }}
    .card-ok button {{ background: {COLORS['OK']['bg']} !important; color: {COLORS['OK']['fg']} !important; border-color: {COLORS['OK']['border']}88 !important; }}
    .card-maior button {{ background: {COLORS['ATUA MAIOR']['bg']} !important; color: {COLORS['ATUA MAIOR']['fg']} !important; border-color: {COLORS['ATUA MAIOR']['border']}88 !important; }}
    .card-menor button {{ background: {COLORS['ATUA MENOR']['bg']} !important; color: {COLORS['ATUA MENOR']['fg']} !important; border-color: {COLORS['ATUA MENOR']['border']}88 !important; }}
    .card-ne button {{ background: {COLORS['NAO ENCONTRADO']['bg']} !important; color: {COLORS['NAO ENCONTRADO']['fg']} !important; border-color: {COLORS['NAO ENCONTRADO']['border']}88 !important; }}
    .card-aberto button {{ background: {COLORS['SALDO ABERTO']['bg']} !important; color: {COLORS['SALDO ABERTO']['fg']} !important; border-color: {COLORS['SALDO ABERTO']['border']}88 !important; }}
    .card-active button {{ border-width: 2.5px !important; box-shadow: 0 0 0 3px rgba(255,255,255,0.12) !important; }}

    /* Hint de filtro ativo */
    .filter-hint {{
        background: #1A2A3A;
        border-left: 3px solid #378ADD;
        padding: 8px 12px;
        border-radius: 4px;
        font-size: 12px;
        color: #CFDFEF;
        margin-bottom: 12px;
    }}
</style>
""", unsafe_allow_html=True)


# ============================================================
# Cabeçalho
# ============================================================
# Logo Pianetto no topo (se disponível)
if LOGO_PIANETTO:
    st.markdown(
        f'<div style="display:flex;align-items:center;justify-content:center;'
        f'padding:12px 0 20px 0;">'
        f'<img src="{LOGO_PIANETTO}" style="max-height:80px;max-width:320px;object-fit:contain;">'
        f'</div>',
        unsafe_allow_html=True,
    )

st.markdown("# Conciliação MOTZ consolidada")
st.caption("PDFs Repom × MOTZ (XLSX) × Cobrança ATUA (XLS) — cruzamento automático")


# ============================================================
# Utilitários
# ============================================================
def parse_rs(v):
    if v is None or (isinstance(v, float) and pd.isna(v)) or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("R$", "").replace("\xa0", " ").strip()
    neg = s.startswith("-") or s.startswith("−")
    s = s.lstrip("-−").strip().replace(".", "").replace(",", ".")
    try:
        n = float(s)
        return -n if neg else n
    except Exception:
        return 0.0


def parse_date_br(v):
    if v is None or v == "" or pd.isna(v):
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    if s in ("", "01/01/0001"):
        return None
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    return None


def fmt_mi(n):
    if n is None:
        return "—"
    abs_n = abs(n)
    sig = "-" if n < 0 else ""
    if abs_n >= 1e6:
        return f"{sig}R$ {abs_n/1e6:,.2f} mi".replace(",", "X").replace(".", ",").replace("X", ".")
    if abs_n >= 1e3:
        return f"{sig}R$ {abs_n/1e3:,.2f} mil".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{sig}R$ {abs_n:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def fmt_rs(n):
    if n is None:
        return "—"
    return f"R$ {n:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def colorir_linhas_tabela(df_pd):
    """
    Colore CÉLULA POR CÉLULA (igual à planilha MOTZ):
    - 'Status' e 'Diferença MOTZ×ATUA': cor do status (verde/vermelho/azul/amarelo)
    - 'Vlr. Saldo': cor do status (como na planilha)
    - 'Situação Saldo': ROXO se "Aberto"
    - 'Diverg. Interna (Quebra/descontos) MOTZ': VERMELHO se diferente de zero
    - Outras colunas: sem fundo
    """
    def _pintar(coluna, row):
        status = str(row.get("Status", "")).strip().upper()
        sit_saldo = str(row.get("Situação Saldo", "")).strip().lower()
        saldo_aberto = "aberto" in sit_saldo

        # Definir cor do status atual
        if status == "OK":
            c_status = COLORS["OK"]
        elif status == "ATUA MAIOR":
            diff = abs(row.get("Diferença MOTZ×ATUA") or 0)
            c_status = COLORS["ATUA MAIOR"] if diff > 100 else COLORS["ATUA MENOR"]
        elif status == "ATUA MENOR":
            c_status = COLORS["ATUA MENOR"]
        elif "ENCONTRADO" in status:
            c_status = COLORS["NAO ENCONTRADO"]
        else:
            c_status = None

        # Coluna "Situação Saldo" → roxo quando Aberto
        if coluna == "Situação Saldo" and saldo_aberto:
            c = COLORS["SALDO ABERTO"]
            return f"background-color: {c['bg']}; color: {c['fg']}; font-weight: 500;"

        # Coluna "Diverg. Interna (Quebra/descontos) MOTZ" → vermelho se diferente de zero
        if coluna == "Diverg. Interna (Quebra/descontos) MOTZ":
            valor_diverg = row.get("Diverg. Interna (Quebra/descontos) MOTZ")
            if valor_diverg is not None and pd.notna(valor_diverg):
                try:
                    if abs(float(valor_diverg)) > 0.01:  # diferente de zero (tolerância 1 centavo)
                        c = COLORS["ATUA MAIOR"]  # usa a mesma paleta vermelha
                        return f"background-color: {c['bg']}; color: {c['fg']}; font-weight: 500;"
                except (ValueError, TypeError):
                    pass

        # Colunas que recebem a cor do status
        if c_status and coluna in ("Status", "Diferença MOTZ×ATUA", "Vlr. Saldo"):
            return f"background-color: {c_status['bg']}; color: {c_status['fg']}; font-weight: 500;"

        return ""

    def _estilo(row):
        return [_pintar(col, row) for col in row.index]

    return df_pd.style.apply(_estilo, axis=1)


# ============================================================
# Executar a skill de conciliação
# ============================================================
def rodar_conciliacao(pdfs_bytes, motz_bytes, atua_bytes, motz_name, atua_name):
    """
    Roda o script scripts/conciliacao.py em um diretório temporário.
    Retorna o caminho do XLSX gerado.
    """
    script_path = Path(__file__).parent / "scripts" / "conciliacao.py"
    if not script_path.exists():
        raise FileNotFoundError(
            "scripts/conciliacao.py não encontrado. Copie o script da skill "
            "/mnt/skills/user/conciliacao-motz/scripts/ para esta pasta."
        )

    tmpdir = tempfile.mkdtemp(prefix="motz_")
    try:
        uploads = Path(tmpdir) / "uploads"
        uploads.mkdir()

        motz_path = uploads / motz_name
        motz_path.write_bytes(motz_bytes)

        atua_path = uploads / atua_name
        atua_path.write_bytes(atua_bytes)

        pdf_paths = []
        seen_hashes = set()
        for name, data in pdfs_bytes:
            h = hashlib.md5(data).hexdigest()
            if h in seen_hashes:
                continue
            seen_hashes.add(h)
            p = uploads / name
            p.write_bytes(data)
            pdf_paths.append(str(p))

        output_path = Path(tmpdir) / "conciliacao_final.xlsx"

        cmd = [
            sys.executable, str(script_path),
            "--motz", str(motz_path),
            "--atua", str(atua_path),
            "--pdfs", *pdf_paths,
            "--output", str(output_path),
        ]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=300)
        if result.returncode != 0:
            raise RuntimeError(
                f"Erro ao rodar conciliação:\n\nSTDOUT:\n{result.stdout}\n\nSTDERR:\n{result.stderr}"
            )

        if not output_path.exists():
            raise RuntimeError(
                f"Script executou mas não gerou o arquivo.\n\n{result.stdout}\n{result.stderr}"
            )

        xlsx_bytes = output_path.read_bytes()
        return xlsx_bytes, result.stdout

    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


# ============================================================
# Processar XLSX → 1 linha por transferência (preserva planilha original)
# ============================================================
# Ordem e nomes oficiais das 22 colunas (espelha a planilha da skill MOTZ)
COLUNAS_OFICIAIS = [
    "Cliente",
    "Contrato",
    "TITULO (NFe)",
    "nr_ctrc ATUA",
    "Nº Carta Frete",
    "Motorista",
    "Nº Romaneio",
    "Data Emissão",
    "Vlr. Frete Líquido",
    "Vlr. Adiantamento",
    "Vlr. Saldo",
    "Soma Adto+Saldo",
    "vl_quebra_avaria",
    "Diverg. Interna (Quebra/descontos) MOTZ",
    "vl_total ATUA",
    "Diferença MOTZ×ATUA",
    "Status",
    "Data Emissão Repom",
    "Data Transferência",
    "Valor Transferido",
    "Situação Adto",
    "Situação Saldo",
]

# Colunas numéricas (formato R$)
COLUNAS_VALOR = {
    "Vlr. Frete Líquido", "Vlr. Adiantamento", "Vlr. Saldo", "Soma Adto+Saldo",
    "vl_quebra_avaria", "Diverg. Interna (Quebra/descontos) MOTZ",
    "vl_total ATUA", "Diferença MOTZ×ATUA", "Valor Transferido",
}

# Colunas de data (formato dd/mm/aaaa)
COLUNAS_DATA = {"Data Emissão", "Data Emissão Repom", "Data Transferência"}


def processar_xlsx(xlsx_bytes):
    """Lê o XLSX de conciliação preservando 1 linha por transferência (igual à planilha original)."""
    xl = pd.ExcelFile(io.BytesIO(xlsx_bytes))
    sheet = next((s for s in xl.sheet_names if "concilia" in s.lower()), xl.sheet_names[0])
    df = pd.read_excel(io.BytesIO(xlsx_bytes), sheet_name=sheet)

    import re as _re
    def find_col(patterns):
        for pat in patterns:
            for c in df.columns:
                if _re.search(pat, str(c), _re.IGNORECASE):
                    return c
        return None

    # Mapeia cada coluna da planilha de origem para o nome oficial
    MAPA = {
        "Cliente":                                  find_col([r"^Cliente"]),
        "Contrato":                                 find_col([r"^Contrato"]),
        "TITULO (NFe)":                             find_col([r"TITULO.*NFe", r"^TITULO", r"^NFe"]),
        "nr_ctrc ATUA":                             find_col([r"nr_ctrc.*ATUA", r"^nr_ctrc", r"^CTRC$"]),
        "Nº Carta Frete":                           find_col([r"N.*Carta.*Frete", r"Carta.Frete"]),
        "Motorista":                                find_col([r"^Motorista"]),
        "Nº Romaneio":                              find_col([r"N.*Romaneio", r"^Romaneio"]),
        "Data Emissão":                             find_col([r"Data Emiss[aã]o$", r"^Data Emiss[aã]o[^R]*$"]),
        "Vlr. Frete Líquido":                       find_col([r"Vlr.*Frete.*L[ií]quido", r"Frete L[ií]quido"]),
        "Vlr. Adiantamento":                        find_col([r"Vlr.*Adiantamento", r"^Adiantamento"]),
        "Vlr. Saldo":                               find_col([r"Vlr\. Saldo", r"^Saldo$"]),
        "Soma Adto+Saldo":                          find_col([r"Soma.*Adto.*Saldo", r"Adto.*Saldo"]),
        "vl_quebra_avaria":                         find_col([r"vl_quebra.avaria", r"quebra.*avaria"]),
        "Diverg. Interna (Quebra/descontos) MOTZ":  find_col([r"Diverg.*Interna", r"Diverg.*MOTZ"]),
        "vl_total ATUA":                            find_col([r"vl_total.*ATUA", r"vl_total"]),
        "Diferença MOTZ×ATUA":                      find_col([r"Diferen.a.*MOTZ.*ATUA", r"Diferen.a.*ATUA"]),
        "Status":                                   find_col([r"^Status"]),
        "Data Emissão Repom":                       find_col([r"Data.*Emiss.*Repom", r"Emiss.*Repom"]),
        "Data Transferência":                       find_col([r"Data.*Transfer", r"Transfer[êe]ncia"]),
        "Valor Transferido":                        find_col([r"Valor Transferido", r"Vlr.*Transferido"]),
        "Situação Adto":                            find_col([r"Situa..o Adto", r"Situa.*Adto"]),
        "Situação Saldo":                           find_col([r"Situa..o Saldo", r"Situa.*Saldo"]),
    }

    if not MAPA["Contrato"] or not MAPA["Status"]:
        raise ValueError(
            f"Planilha não reconhecida. Colunas encontradas: {list(df.columns)}"
        )

    # Montar DataFrame de saída preservando 1 linha por transferência
    linhas = []
    for _, row in df.iterrows():
        contrato = str(row[MAPA["Contrato"]]).strip() if pd.notna(row[MAPA["Contrato"]]) else ""
        if not contrato or contrato == "nan":
            continue

        nova = {}
        for col_oficial, col_origem in MAPA.items():
            if col_origem is None:
                # Coluna não encontrada na planilha de origem → deixa vazio/None
                nova[col_oficial] = None if col_oficial in COLUNAS_VALOR or col_oficial in COLUNAS_DATA else ""
                continue

            val = row[col_origem]

            if col_oficial in COLUNAS_VALOR:
                nova[col_oficial] = parse_rs(val) if pd.notna(val) and str(val).strip() not in ("", "nan") else None
            elif col_oficial in COLUNAS_DATA:
                nova[col_oficial] = parse_date_br(val)
            else:
                nova[col_oficial] = str(val).strip() if pd.notna(val) else ""

        linhas.append(nova)

    df_out = pd.DataFrame(linhas)

    # Garantir ordem oficial das colunas
    for col in COLUNAS_OFICIAIS:
        if col not in df_out.columns:
            df_out[col] = None
    df_out = df_out[COLUNAS_OFICIAIS]

    return df_out


# ============================================================
# MESCLAGEM INTELIGENTE (dedup por contrato+valor+data)
# ============================================================
def _chave_linha(row):
    """Chave única por transferência: Contrato + Valor Transferido + Data Transferência."""
    c = str(row.get("Contrato", "")).strip()
    v = row.get("Valor Transferido")
    v_str = f"{float(v):.2f}" if pd.notna(v) else "0.00"
    d = row.get("Data Transferência")
    if pd.notna(d) and hasattr(d, "strftime"):
        d_str = d.strftime("%Y-%m-%d")
    else:
        d_str = ""
    return f"{c}|{v_str}|{d_str}"


def mesclar_dataframes(df_base, df_novo):
    """Mescla df_novo em df_base deduplicando. Retorna (df, stats)."""
    if df_base is None or len(df_base) == 0:
        return df_novo.copy(), {"novas": len(df_novo), "duplicadas": 0, "total": len(df_novo)}
    if df_novo is None or len(df_novo) == 0:
        return df_base.copy(), {"novas": 0, "duplicadas": 0, "total": len(df_base)}

    chaves_base = set(df_base.apply(_chave_linha, axis=1))
    df_novo = df_novo.copy()
    df_novo["_chave"] = df_novo.apply(_chave_linha, axis=1)
    mask_novas = ~df_novo["_chave"].isin(chaves_base)
    df_inserir = df_novo[mask_novas].drop(columns=["_chave"])
    duplicadas = int((~mask_novas).sum())

    df_final = pd.concat([df_base, df_inserir], ignore_index=True)
    return df_final, {"novas": len(df_inserir), "duplicadas": duplicadas, "total": len(df_final)}


def gerar_xlsx_historico(df):
    """Gera XLSX colorido com as 22 colunas, pronto para baixar/compartilhar."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "conciliacao"

    df_write = df.copy()
    for col in COLUNAS_DATA:
        if col in df_write.columns:
            df_write[col] = df_write[col].apply(
                lambda d: d.strftime("%d/%m/%Y") if pd.notna(d) and hasattr(d, "strftime") else ""
            )

    for row in dataframe_to_rows(df_write, index=False, header=True):
        ws.append(row)

    # Cores claras (boa visibilidade no Excel)
    fills = {
        "OK":     PatternFill("solid", fgColor="D5E8C1"),
        "MAIOR":  PatternFill("solid", fgColor="F8CCCC"),
        "MENOR":  PatternFill("solid", fgColor="CDE3F7"),
        "NE":     PatternFill("solid", fgColor="FCE9B6"),
        "ABERTO": PatternFill("solid", fgColor="DDD9FB"),
    }

    header = [cell.value for cell in ws[1]]
    def idx(nome):
        try:
            return header.index(nome) + 1
        except ValueError:
            return None

    col_status = idx("Status")
    col_diff = idx("Diferença MOTZ×ATUA")
    col_saldo = idx("Vlr. Saldo")
    col_sit_saldo = idx("Situação Saldo")

    for row_idx in range(2, ws.max_row + 1):
        status_val = ws.cell(row=row_idx, column=col_status).value if col_status else ""
        status = str(status_val).strip().upper() if status_val else ""
        sit_saldo_val = ws.cell(row=row_idx, column=col_sit_saldo).value if col_sit_saldo else ""
        saldo_aberto = "aberto" in str(sit_saldo_val).lower() if sit_saldo_val else False

        if status == "OK":
            fill = fills["OK"]
        elif status == "ATUA MAIOR":
            diff_val = ws.cell(row=row_idx, column=col_diff).value if col_diff else 0
            try:
                diff_abs = abs(float(diff_val)) if diff_val else 0
            except (ValueError, TypeError):
                diff_abs = 0
            fill = fills["MAIOR"] if diff_abs > 100 else fills["MENOR"]
        elif status == "ATUA MENOR":
            fill = fills["MENOR"]
        elif "ENCONTRADO" in status:
            fill = fills["NE"]
        else:
            fill = None

        if fill:
            for c in (col_status, col_diff, col_saldo):
                if c:
                    ws.cell(row=row_idx, column=c).fill = fill

        if saldo_aberto and col_sit_saldo:
            ws.cell(row=row_idx, column=col_sit_saldo).fill = fills["ABERTO"]

    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
        cell.fill = PatternFill("solid", fgColor="E8E8E8")

    for col in ws.columns:
        try:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)
        except Exception:
            pass

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# UPLOAD
# ============================================================
# BASE HISTÓRICA (topo — aparece sempre)
# ============================================================
with st.container(border=True):
    col_h1, col_h2 = st.columns([2, 1])
    with col_h1:
        st.markdown("### 📚 Base Histórica")
        if "df" in st.session_state and len(st.session_state["df"]) > 0:
            df_atual = st.session_state["df"]
            n_linhas = len(df_atual)
            n_contratos = df_atual["Contrato"].nunique()
            st.caption(f"✅ **{n_linhas} transferências** de **{n_contratos} contratos** carregados na base")
        else:
            st.caption("Nenhuma base carregada. Comece subindo uma base histórica existente ou rodando uma nova conciliação abaixo.")

    with col_h2:
        if "df" in st.session_state and len(st.session_state["df"]) > 0:
            try:
                xlsx_base = gerar_xlsx_historico(st.session_state["df"])
                st.download_button(
                    "💾 Baixar Base Histórica",
                    data=xlsx_base,
                    file_name=f"base_historica_conciliacao_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary",
                )
            except Exception as e:
                st.caption(f"Erro ao gerar XLSX: {e}")

    col_h3, col_h4 = st.columns([1, 1])
    with col_h3:
        with st.expander("📤 Carregar base histórica existente"):
            st.caption("Sobrescreve a base atual. Use essa opção no início do dia para carregar a base do time.")
            xlsx_base_upload = st.file_uploader(
                "Base histórica (XLSX)",
                type=["xlsx"],
                key="base_historica_upload",
                label_visibility="collapsed",
            )
            if xlsx_base_upload is not None and st.session_state.get("last_base_loaded") != xlsx_base_upload.name:
                try:
                    df_hist = processar_xlsx(xlsx_base_upload.read())
                    st.session_state["df"] = df_hist
                    st.session_state["origem"] = f"Base histórica: {xlsx_base_upload.name}"
                    st.session_state["last_base_loaded"] = xlsx_base_upload.name
                    st.success(f"✓ Base carregada: {len(df_hist)} transferências")
                    st.rerun()
                except Exception as e:
                    st.error(f"Erro ao carregar: {e}")

    with col_h4:
        with st.expander("🗑️ Limpar base"):
            st.caption("⚠️ Apaga tudo. Ação irreversível (mas você pode recarregar o XLSX).")
            if st.button("🗑️ Limpar tudo e recomeçar", use_container_width=True):
                for k in ["df", "xlsx_bytes", "origem", "status_click", "last_base_loaded", "modo_xlsx_pronto"]:
                    if k in st.session_state:
                        del st.session_state[k]
                st.success("Base limpa!")
                st.rerun()


# ============================================================
# UPLOAD DE NOVOS ARQUIVOS (para MESCLAR na base)
# ============================================================
with st.container(border=True):
    st.markdown("### 📂 Adicionar novos arquivos")
    base_msg = "**Mesclar** com a base existente" if "df" in st.session_state else "**Começar uma nova base**"
    st.caption(
        f"Suba PDFs Repom + MOTZ + ATUA para {base_msg}. "
        f"Linhas duplicadas (mesmo contrato + mesmo valor + mesma data) serão ignoradas automaticamente."
    )

    col1, col2, col3 = st.columns(3)

    with col1:
        if LOGO_REPOM:
            st.markdown(
                f'<div style="display:flex;align-items:center;justify-content:center;'
                f'height:60px;background:#FFFFFF;border-radius:8px;padding:6px;margin-bottom:8px;">'
                f'<img src="{LOGO_REPOM}" style="max-height:48px;max-width:100%;object-fit:contain;"></div>',
                unsafe_allow_html=True,
            )
        st.markdown('<div style="text-align:center;font-weight:500;margin-bottom:4px;">PDFs Repom</div>', unsafe_allow_html=True)
        pdfs = st.file_uploader(
            "Transferências bancárias",
            type=["pdf"],
            accept_multiple_files=True,
            key="pdfs",
            label_visibility="collapsed",
        )
        if pdfs:
            st.caption(f"✓ {len(pdfs)} PDF(s)")

    with col2:
        if LOGO_MOTZ:
            st.markdown(
                f'<div style="display:flex;align-items:center;justify-content:center;'
                f'height:60px;background:#FFFFFF;border-radius:8px;padding:6px;margin-bottom:8px;">'
                f'<img src="{LOGO_MOTZ}" style="max-height:48px;max-width:100%;object-fit:contain;"></div>',
                unsafe_allow_html=True,
            )
        st.markdown('<div style="text-align:center;font-weight:500;margin-bottom:4px;">Arquivo MOTZ</div>', unsafe_allow_html=True)
        motz = st.file_uploader(
            "export*.xlsx",
            type=["xlsx"],
            key="motz",
            label_visibility="collapsed",
        )
        if motz:
            st.caption(f"✓ {motz.name}")

    with col3:
        if LOGO_ATUA:
            st.markdown(
                f'<div style="display:flex;align-items:center;justify-content:center;'
                f'height:60px;background:#FFFFFF;border-radius:8px;padding:6px;margin-bottom:8px;">'
                f'<img src="{LOGO_ATUA}" style="max-height:48px;max-width:100%;object-fit:contain;"></div>',
                unsafe_allow_html=True,
            )
        st.markdown('<div style="text-align:center;font-weight:500;margin-bottom:4px;">Cobrança ATUA</div>', unsafe_allow_html=True)
        atua = st.file_uploader(
            "*cobranca*.xls",
            type=["xls", "xlsx"],
            key="atua",
            label_visibility="collapsed",
        )
        if atua:
            st.caption(f"✓ {atua.name}")

    col_b1, _ = st.columns([1, 3])
    with col_b1:
        rodar_btn = st.button("🔄 Processar e mesclar", type="primary", use_container_width=True, disabled=not (pdfs and motz and atua))


# ============================================================
# (bloco "Carregar XLSX pronto" agora está integrado ao topo)
# ============================================================


# ============================================================
# Rodar conciliação
# ============================================================
if rodar_btn and pdfs and motz and atua:
    with st.spinner("Rodando conciliação... isso pode levar 30s-2min dependendo do tamanho dos arquivos."):
        try:
            pdfs_data = [(f.name, f.read()) for f in pdfs]
            motz_data = motz.read()
            atua_data = atua.read()

            xlsx_bytes, log = rodar_conciliacao(
                pdfs_data, motz_data, atua_data, motz.name, atua.name
            )
            df_novo = processar_xlsx(xlsx_bytes)

            # MESCLAR com a base existente (deduplicação inteligente)
            df_base_atual = st.session_state.get("df")
            df_final, stats = mesclar_dataframes(df_base_atual, df_novo)

            st.session_state["df"] = df_final
            st.session_state["xlsx_bytes"] = xlsx_bytes  # último processamento bruto
            st.session_state["origem"] = (
                f"Mesclado às {datetime.now().strftime('%H:%M:%S')} · "
                f"{len(pdfs_data)} PDFs + {motz.name} + {atua.name}"
            )
            st.session_state["log"] = log

            # Relatório de mesclagem
            if df_base_atual is None or len(df_base_atual) == 0:
                st.success(f"✓ Base inicial criada · {stats['total']} transferências")
            else:
                msg_partes = [f"✓ Mesclagem concluída!"]
                if stats["novas"] > 0:
                    msg_partes.append(f"📥 **{stats['novas']} novas** transferências adicionadas")
                if stats["duplicadas"] > 0:
                    msg_partes.append(f"♻️ **{stats['duplicadas']} duplicadas** ignoradas (já existiam)")
                msg_partes.append(f"📊 **Total agora:** {stats['total']} transferências")
                st.success("  ·  ".join(msg_partes))

            st.info("💡 Não esqueça de **baixar a base histórica atualizada** no topo da página e compartilhar com o time!", icon="💾")

        except Exception as e:
            st.error(f"Erro na conciliação:\n\n{str(e)}")
            st.stop()


# ============================================================
# Dashboard (quando houver dados)
# ============================================================
if "df" in st.session_state:
    df = st.session_state["df"]
    st.divider()

    # Inicializar filtro clicável no session_state
    if "status_click" not in st.session_state:
        st.session_state["status_click"] = None  # None = sem filtro

    # Cabeçalho do dashboard
    col_a, col_b = st.columns([3, 1])
    with col_a:
        st.caption(f"🟢 {st.session_state.get('origem', '')}")
    with col_b:
        if "xlsx_bytes" in st.session_state:
            st.download_button(
                "⬇️ XLSX do último processamento",
                data=st.session_state["xlsx_bytes"],
                file_name=f"ultimo_processamento_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                help="XLSX da última rodada (apenas os arquivos processados agora). Para a base histórica completa, use o botão no topo.",
            )

    st.info(
        "**1 linha por transferência Repom** — igual à planilha da skill. "
        "Se um contrato teve 2 transferências (ex: adto + saldo), aparece 2 vezes. "
        "As colunas coloridas destacam divergências e saldo em aberto.",
        icon="ℹ️",
    )

    # Filtros
    with st.container(border=True):
        datas_validas = df["Data Emissão"].dropna()
        if len(datas_validas) > 0:
            date_min = datas_validas.min().date()
            date_max = datas_validas.max().date()
        else:
            date_min = date_max = datetime.now().date()

        # Folga de ±1 ano para o usuário poder escolher qualquer data
        from datetime import date as _date
        date_min_widget = _date(date_min.year - 1, 1, 1)
        date_max_widget = _date(date_max.year + 1, 12, 31)

        col_f1, col_f2, col_f3, col_f4 = st.columns([1, 1, 1, 2])
        with col_f1:
            date_from = st.date_input(
                "De",
                value=date_min,
                min_value=date_min_widget,
                max_value=date_max_widget,
                format="DD/MM/YYYY",
                key="date_from_input",
            )
        with col_f2:
            date_to = st.date_input(
                "Até",
                value=date_max,
                min_value=date_min_widget,
                max_value=date_max_widget,
                format="DD/MM/YYYY",
                key="date_to_input",
            )
        with col_f3:
            status_filter = st.selectbox(
                "Status",
                ["Todos", "OK", "ATUA MAIOR", "ATUA MENOR", "NÃO ENCONTRADO", "Saldo aberto"],
                key="status_dropdown",
            )
        with col_f4:
            busca = st.text_input(
                "Buscar",
                placeholder="contrato, CTRC, motorista, NFe...",
                key="busca_input",
            )

    # Aplicar filtros de data
    df_f = df.copy()
    if date_from:
        df_f = df_f[df_f["Data Emissão"].apply(lambda d: pd.isna(d) or d.date() >= date_from)]
    if date_to:
        df_f = df_f[df_f["Data Emissão"].apply(lambda d: pd.isna(d) or d.date() <= date_to)]

    df_periodo = df_f.copy()  # para KPIs do período (sem filtro de status)

    # ============================================================
    # Combinar filtro do dropdown + filtro clicável
    # Regra: se o usuário clicou num card/fatia, isso tem prioridade sobre o dropdown "Todos"
    # Se o dropdown ≠ "Todos" E clicou num card, o clique vence (sincroniza)
    # ============================================================
    filtro_ativo = st.session_state.get("status_click")
    if status_filter != "Todos":
        filtro_ativo = status_filter  # dropdown vence se não for "Todos"

    if filtro_ativo == "Saldo aberto":
        df_f = df_f[df_f["Situação Saldo"] == "Aberto"]
    elif filtro_ativo and filtro_ativo != "Todos":
        df_f = df_f[df_f["Status"] == filtro_ativo]

    if busca:
        b = busca.lower()
        mask = (
            df_f["Contrato"].astype(str).str.lower().str.contains(b, na=False) |
            df_f["nr_ctrc ATUA"].astype(str).str.lower().str.contains(b, na=False) |
            df_f["Motorista"].astype(str).str.lower().str.contains(b, na=False) |
            df_f["TITULO (NFe)"].astype(str).str.lower().str.contains(b, na=False) |
            df_f["Cliente"].astype(str).str.lower().str.contains(b, na=False) |
            df_f["Nº Carta Frete"].astype(str).str.lower().str.contains(b, na=False) |
            df_f["Nº Romaneio"].astype(str).str.lower().str.contains(b, na=False)
        )
        df_f = df_f[mask]

    # ============================================================
    # KPIs (sobre df_periodo, sem filtro de status)
    # Deduplicar por contrato para somas (Frete/ATUA se repetem em 2+ transferências)
    # ============================================================
    total_linhas = len(df_periodo)
    df_unicos = df_periodo.drop_duplicates(subset=["Contrato"]) if total_linhas > 0 else df_periodo
    total = len(df_unicos)  # nº de contratos únicos

    ok_n = (df_unicos["Status"] == "OK").sum()
    maior_n = (df_unicos["Status"] == "ATUA MAIOR").sum()
    menor_n = (df_unicos["Status"] == "ATUA MENOR").sum()
    ne_n = (df_unicos["Status"] == "NÃO ENCONTRADO").sum()
    aberto_n = (df_unicos["Situação Saldo"] == "Aberto").sum()

    soma_motz = df_unicos["Vlr. Frete Líquido"].fillna(0).sum()
    soma_atua = df_unicos["vl_total ATUA"].fillna(0).sum()
    # Valor Transferido: SOMA TODAS as linhas (cada linha = 1 transferência)
    soma_transf = df_periodo["Valor Transferido"].fillna(0).sum()
    contratos_com_transf = df_periodo[df_periodo["Valor Transferido"].fillna(0) > 0]["Contrato"].nunique()
    soma_saldo_aberto = df_unicos[df_unicos["Situação Saldo"] == "Aberto"]["Vlr. Saldo"].fillna(0).sum()

    indice = (ok_n / total * 100) if total else 0

    col_k1, col_k2, col_k3, col_k4, col_k5 = st.columns(5)
    with col_k1:
        st.metric("Índice conciliação", f"{indice:.1f}%".replace(".", ","), f"{ok_n} de {total} OK")
    with col_k2:
        st.metric("Soma MOTZ", fmt_mi(soma_motz), help="Frete líquido (sem duplicar por transferência)")
    with col_k3:
        diff = soma_atua - soma_motz
        st.metric("Soma ATUA", fmt_mi(soma_atua), delta=fmt_mi(diff), delta_color="inverse")
    with col_k4:
        st.metric("Transferido Repom", fmt_mi(soma_transf), f"{contratos_com_transf} contratos c/ PDF")
    with col_k5:
        st.metric("Saldo em aberto", fmt_mi(soma_saldo_aberto), f"{aberto_n} contratos")

    # ============================================================
    # Distribuição clicável (cards + gráfico de pizza)
    # ============================================================
    st.markdown("### Distribuição por status")
    st.caption("👆 Clique em um card ou numa fatia do gráfico para filtrar a tabela. Clique de novo para limpar.")

    def pct(n):
        return f"{n/total*100:.1f}%".replace(".", ",") if total else "0,0%"

    # ---------- Cards clicáveis (5 colunas) ----------
    cards = [
        ("OK",             ok_n,     "card-ok",      "🟢"),
        ("ATUA MAIOR",     maior_n,  "card-maior",   "🔴"),
        ("ATUA MENOR",     menor_n,  "card-menor",   "🔵"),
        ("NÃO ENCONTRADO", ne_n,     "card-ne",      "🟡"),
        ("Saldo aberto",   aberto_n, "card-aberto",  "🟣"),
    ]

    cols_cards = st.columns(5)
    filtro_atual_clique = st.session_state.get("status_click")

    for idx, (label, n, css_class, emoji) in enumerate(cards):
        with cols_cards[idx]:
            ativo = (filtro_atual_clique == label)
            classe_final = f"{css_class} card-active" if ativo else css_class
            st.markdown(f'<div class="{classe_final}">', unsafe_allow_html=True)
            if st.button(
                f"{emoji}  **{label}**\n\n{n} · {pct(n)}",
                key=f"card_{label}",
                use_container_width=True,
            ):
                # Toggle: se já tá selecionado, desseleciona
                if st.session_state.get("status_click") == label:
                    st.session_state["status_click"] = None
                else:
                    st.session_state["status_click"] = label
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

    # Aviso de filtro ativo
    if filtro_atual_clique and status_filter == "Todos":
        st.markdown(
            f'<div class="filter-hint">🎯 Filtro ativo pelo card: <b>{filtro_atual_clique}</b> '
            f'· Exibindo {len(df_f)} de {total} contratos. Clique no mesmo card para limpar.</div>',
            unsafe_allow_html=True,
        )

    # ---------- Gráfico de pizza (abaixo dos cards) ----------
    col_g1, col_g2 = st.columns([1, 1])

    with col_g1:
        st.markdown("**Distribuição visual**")
        dist_data = pd.DataFrame([
            {"Status": "OK",             "Qtd": ok_n,    "Cor": COLORS["OK"]["plot"]},
            {"Status": "ATUA MAIOR",     "Qtd": maior_n, "Cor": COLORS["ATUA MAIOR"]["plot"]},
            {"Status": "ATUA MENOR",     "Qtd": menor_n, "Cor": COLORS["ATUA MENOR"]["plot"]},
            {"Status": "NÃO ENCONTRADO", "Qtd": ne_n,    "Cor": COLORS["NAO ENCONTRADO"]["plot"]},
            {"Status": "Saldo aberto",   "Qtd": aberto_n,"Cor": COLORS["SALDO ABERTO"]["plot"]},
        ])
        dist_data = dist_data[dist_data["Qtd"] > 0]  # não plota fatias vazias

        if len(dist_data) > 0:
            fig = px.pie(
                dist_data,
                values="Qtd",
                names="Status",
                color="Status",
                color_discrete_map={r["Status"]: r["Cor"] for _, r in dist_data.iterrows()},
                hole=0.4,
            )
            fig.update_traces(
                textposition="inside",
                textinfo="percent+label",
                hovertemplate="<b>%{label}</b><br>%{value} contratos<br>%{percent}<extra></extra>",
                pull=[0.08 if s == filtro_atual_clique else 0 for s in dist_data["Status"]],
            )
            fig.update_layout(
                height=280,
                margin=dict(t=10, b=10, l=10, r=10),
                showlegend=True,
                legend=dict(orientation="v", yanchor="middle", y=0.5, xanchor="left", x=1.05, font=dict(size=11, color="#EAEAEA")),
                paper_bgcolor="#1A1A1A",
                plot_bgcolor="#1A1A1A",
                font=dict(color="#EAEAEA"),
            )
            # Eventos de clique no gráfico
            evento = st.plotly_chart(fig, use_container_width=True, on_select="rerun", key="pie_chart")
            if evento and evento.get("selection") and evento["selection"].get("points"):
                ponto = evento["selection"]["points"][0]
                status_clicado = ponto.get("label")
                if status_clicado:
                    # Toggle
                    if st.session_state.get("status_click") == status_clicado:
                        st.session_state["status_click"] = None
                    else:
                        st.session_state["status_click"] = status_clicado
                    st.rerun()
        else:
            st.caption("Sem dados para plotar")

    with col_g2:
        st.markdown("**Frete líquido emitido por dia**")
        df_chart = df_unicos.dropna(subset=["Data Emissão"]).copy()
        if len(df_chart) > 0:
            df_chart["Dia"] = df_chart["Data Emissão"].dt.date
            daily = df_chart.groupby("Dia")["Vlr. Frete Líquido"].sum().reset_index()
            daily.columns = ["Data", "Frete Líquido"]
            st.bar_chart(daily, x="Data", y="Frete Líquido", height=280, color="#378ADD")
        else:
            st.caption("Sem dados de data para o período")

    # ============================================================
    # Tabela (com cores por linha)
    # ============================================================
    st.markdown(f"**Transferências · {len(df_f)} linhas exibidas** " +
                (f"(filtro: {filtro_ativo})" if filtro_ativo and filtro_ativo != "Todos" else ""))

    # ---- Multi-select para escolher quais colunas exibir ----
    # Seleção padrão (mais usadas) — usuário pode expandir
    PADRAO_VISIVEIS = [
        "Cliente", "Contrato", "TITULO (NFe)", "Motorista",
        "Data Emissão", "Vlr. Frete Líquido", "Vlr. Saldo",
        "vl_total ATUA", "Diferença MOTZ×ATUA", "Status",
        "Data Transferência", "Valor Transferido",
        "Situação Adto", "Situação Saldo",
    ]

    # Estratégia: usar chave dinâmica para permitir reset pelos botões
    if "colunas_default" not in st.session_state:
        st.session_state["colunas_default"] = PADRAO_VISIVEIS
    if "colunas_version" not in st.session_state:
        st.session_state["colunas_version"] = 0

    with st.expander("⚙️ Escolher colunas visíveis", expanded=False):
        col_rst1, col_rst2, _ = st.columns([1, 1, 3])
        with col_rst1:
            if st.button("✅ Mostrar todas", key="btn_todas"):
                st.session_state["colunas_default"] = list(COLUNAS_OFICIAIS)
                st.session_state["colunas_version"] += 1
                st.rerun()
        with col_rst2:
            if st.button("↺ Padrão", key="btn_padrao"):
                st.session_state["colunas_default"] = PADRAO_VISIVEIS
                st.session_state["colunas_version"] += 1
                st.rerun()

        colunas_escolhidas = st.multiselect(
            "Selecione as colunas que deseja ver (todas as 22 disponíveis):",
            options=COLUNAS_OFICIAIS,
            default=st.session_state["colunas_default"],
            key=f"colunas_multiselect_v{st.session_state['colunas_version']}",
        )

    if not colunas_escolhidas:
        colunas_escolhidas = PADRAO_VISIVEIS

    # Ordenar colunas escolhidas na ORDEM OFICIAL (não na ordem de seleção)
    colunas_ordenadas = [c for c in COLUNAS_OFICIAIS if c in colunas_escolhidas]

    # Preparar df_show
    df_show = df_f.copy().sort_values(
        ["Data Emissão", "Contrato"],
        ascending=[False, True],
        na_position="last",
    )

    # Construir view da tabela
    df_tabela = df_show[colunas_ordenadas].reset_index(drop=True)

    # Formatadores
    formatadores = {}
    for col in colunas_ordenadas:
        if col in COLUNAS_VALOR:
            formatadores[col] = lambda v: (
                f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                if pd.notna(v) and v != 0 else ("R$ 0,00" if v == 0 else "—")
            )
        elif col in COLUNAS_DATA:
            formatadores[col] = lambda d: d.strftime("%d/%m/%Y") if pd.notna(d) and isinstance(d, datetime) else "—"

    # Aplicar cores + formatação
    styler = colorir_linhas_tabela(df_tabela)
    if formatadores:
        styler = styler.format(formatadores)

    st.dataframe(styler, use_container_width=True, hide_index=True, height=520)

    # Legenda de cores
    with st.expander("🎨 Legenda de cores", expanded=False):
        st.markdown(f"""
        As cores seguem **exatamente** a planilha MOTZ original (célula por célula):

        - <span class="status-ok">🟢 OK</span> — colunas **Status**, **Diferença MOTZ×ATUA** e **Vlr. Saldo** em verde
        - <span class="status-maior">🔴 ATUA MAIOR > R$100</span> — mesmas colunas em vermelho (diferença crítica)
        - <span class="status-menor">🔵 ATUA MAIOR até R$100 / ATUA MENOR</span> — mesmas colunas em azul (diferença pequena)
        - <span class="status-ne">🟡 NÃO ENCONTRADO</span> — mesmas colunas em amarelo
        - <span class="status-aberto">🟣 Situação Saldo = Aberto</span> — apenas a coluna **Situação Saldo** em roxo

        Assim você vê ao mesmo tempo: status da conferência MOTZ×ATUA + pendência de saldo.
        """, unsafe_allow_html=True)

    # Export dos filtrados (todas as colunas, sem depender do que está visível)
    csv = df_f.to_csv(index=False).encode("utf-8")
    st.download_button(
        "⬇️ Baixar tabela filtrada (CSV · todas as colunas)",
        csv,
        f"conciliacao_filtrado_{datetime.now().strftime('%Y-%m-%d')}.csv",
        "text/csv",
    )

else:
    # Tela inicial sem dados
    st.info(
        "👆 **Comece subindo os 3 arquivos** (PDFs Repom, MOTZ XLSX, ATUA XLS) e clique em "
        "**Rodar conciliação**. Ou use **Carregar XLSX pronto** se você já tem a planilha consolidada gerada.",
        icon="📤",
    )

    with st.expander("ℹ️ Sobre esta ferramenta"):
        st.markdown("""
        Este aplicativo executa a skill `conciliacao-motz` diretamente no servidor, fazendo o cruzamento entre três fontes:

        1. **PDFs Repom** — transferências bancárias (chave: Contrato)
        2. **Arquivo MOTZ** — relatório de cartas-frete (chave: Nº formulário = Contrato do PDF)
        3. **Cobrança ATUA** — API Contabilidade (chave: nr_nf = NF cliente do MOTZ)

        **Saídas:**
        - Planilha XLSX com formatação condicional (verde/vermelho/azul/amarelo/roxo)
        - Dashboard interativo com KPIs, filtros, busca e **distribuição clicável**
        - Exportação filtrada em CSV

        **Novidades v3:**
        - 🎨 Cores da skill aplicadas em toda a tabela (verde/vermelho/azul/amarelo/roxo)
        - 👆 Cards de status e gráfico de pizza clicáveis (filtra a tabela)

        **Limite:** 200 MB por arquivo (configurável no Streamlit).
        """)

# Footer
st.divider()
st.caption("Dashboard Conciliação MOTZ · skill conciliacao-motz · Streamlit Cloud · v4.2 (logo Pianetto)")
