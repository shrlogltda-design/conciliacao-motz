#!/usr/bin/env python3
"""
Conciliação Bancária MOTZ TRANSPORTES
Cruza 3 fontes: PDFs Repom, arquivo MOTZ (XLSX), arquivo ATUA (XLS)
Gera planilha Excel final com verificações e cores.

v4.1 - FIX: NFs separadas por virgula no nr_nf do ATUA agora dao match corretamente.
"""
import argparse
import os
import re
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, os.path.dirname(__file__))

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
except ImportError:
    print("ERRO: openpyxl nao encontrado. Instale com: pip install openpyxl")
    sys.exit(1)

try:
    import pdfplumber
except ImportError:
    pdfplumber = None


GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT = Font(color="006100", bold=False)
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(color="9C0006", bold=True)
BLUE_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
BLUE_FONT = Font(color="1F4E79", bold=True)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
PURPLE_FILL = PatternFill(start_color="D5B8EA", end_color="D5B8EA", fill_type="solid")
PURPLE_FONT = Font(color="5B2C6F", bold=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def _split_nf_list(nf_raw):
    """
    Quebra uma string com NFs separadas por virgula/ponto-virgula em lista.
    Exemplo: "17272, 17271" -> ["17272", "17271"]
             "156028,156028" -> ["156028"] (dedup)
             "32672.0" -> ["32672"]
    """
    if nf_raw is None:
        return []
    s = str(nf_raw).strip()
    if not s or s.lower() == "nan":
        return []
    parts = re.split(r'[,;]', s)
    result = []
    seen = set()
    for p in parts:
        p = p.strip()
        if not p:
            continue
        if '.' in p:
            p = p.split('.')[0]
        p_clean = p.lstrip('0') or '0'
        if p_clean not in seen:
            seen.add(p_clean)
            result.append(p_clean)
    return result


def parse_pdf_repom(pdf_paths):
    import hashlib
    transfers = []
    unique_paths = []
    seen_hashes = set()
    for pdf_path in pdf_paths:
        try:
            with open(pdf_path, 'rb') as f:
                file_hash = hashlib.md5(f.read()).hexdigest()
            if file_hash in seen_hashes:
                print(f"  Ignorando arquivo duplicado: {pdf_path}")
                continue
            seen_hashes.add(file_hash)
            unique_paths.append(pdf_path)
        except Exception as e:
            print(f"  AVISO: nao foi possivel ler {pdf_path}: {e}")
            unique_paths.append(pdf_path)

    print(f"  {len(pdf_paths)} PDFs recebidos -> {len(unique_paths)} unicos (por hash)")

    if pdfplumber is None:
        print("  AVISO: pdfplumber nao disponivel. Tentando regex no texto...")
        for pdf_path in unique_paths:
            transfers.extend(_parse_pdf_fallback(pdf_path))
        return transfers

    for pdf_path in unique_paths:
        print(f"  Lendo PDF: {pdf_path}")
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    words = page.extract_words()
                    if words:
                        transfers.extend(_parse_repom_words(words))
                    else:
                        text = page.extract_text()
                        if text:
                            transfers.extend(_parse_repom_text(text))
        except Exception as e:
            print(f"  ERRO ao ler PDF {pdf_path}: {e}")
            transfers.extend(_parse_pdf_fallback(pdf_path))

    return transfers


def _parse_repom_words(words):
    transfers = []
    header_y = None
    col_positions = {}

    for w in words:
        if w['text'] == 'Contrato':
            header_y = w['top']
            break

    if header_y is None:
        return transfers

    header_words = [w for w in words if abs(w['top'] - header_y) < 8]
    for w in header_words:
        txt = w['text'].lower()
        x = w['x0']
        if 'cliente' in txt:
            col_positions['cliente_x'] = x
        elif 'contrato' in txt:
            col_positions['contrato_x'] = x
        elif txt == 'emissao' or 'emiss' in txt:
            col_positions['data_emissao_x'] = x - 30
        elif 'quita' in txt:
            col_positions['data_quitacao_x'] = x - 20
        elif 'pagamento' in txt:
            col_positions['data_pagamento_x'] = x - 20
        elif 'valor' in txt:
            col_positions['valor_x'] = x

    for w in words:
        if w['text'].lower() == 'transferencia' and abs(w['top'] - header_y) < 12:
            col_positions['data_transferencia_x'] = w['x0']

    data_words_header = [w for w in header_words if w['text'].lower() == 'data']
    data_words_header.sort(key=lambda w: w['x0'])
    if len(data_words_header) >= 1:
        col_positions.setdefault('data_emissao_x', data_words_header[0]['x0'])
    if len(data_words_header) >= 2:
        col_positions.setdefault('data_quitacao_x', data_words_header[1]['x0'])
    if len(data_words_header) >= 3:
        col_positions.setdefault('data_pagamento_x', data_words_header[2]['x0'])

    data_words = [w for w in words if w['top'] > header_y + 10]
    footer_keywords = ['total', 'alameda', 'periodo']
    data_words = [w for w in data_words
                  if not any(kw in w['text'].lower() for kw in footer_keywords)]

    if not data_words:
        return transfers

    lines = {}
    for w in data_words:
        y = round(w['top'] / 4) * 4
        lines.setdefault(y, []).append(w)

    sorted_lines = sorted(lines.items())

    current_transfer = None
    contrato_x = col_positions.get('contrato_x', 94)
    valor_x = col_positions.get('valor_x', 400)
    transfer_x = col_positions.get('data_transferencia_x', 333)

    for y, line_words in sorted_lines:
        line_words.sort(key=lambda w: w['x0'])

        for w in line_words:
            txt = w['text']
            x = w['x0']

            if re.match(r'^\d{7,}$', txt) and abs(x - contrato_x) < 30:
                if current_transfer and current_transfer.get('contrato'):
                    transfers.append(current_transfer)
                current_transfer = {
                    'contrato': txt,
                    'cliente': '',
                    'data_emissao': '',
                    'data_quitacao': '',
                    'data_pagamento': '',
                    'data_transferencia': '',
                    'valor': 0.0,
                }
                continue

            if current_transfer is None:
                continue

            if re.match(r'\d{2}/\d{2}/\d{4}$', txt):
                date_cols = {}
                if 'data_emissao_x' in col_positions:
                    date_cols['data_emissao'] = abs(x - col_positions['data_emissao_x'])
                if 'data_quitacao_x' in col_positions:
                    date_cols['data_quitacao'] = abs(x - col_positions['data_quitacao_x'])
                if 'data_pagamento_x' in col_positions:
                    date_cols['data_pagamento'] = abs(x - col_positions['data_pagamento_x'])
                date_cols['data_transferencia'] = abs(x - transfer_x)

                if date_cols:
                    best_col = min(date_cols, key=date_cols.get)
                    if date_cols[best_col] < 40:
                        current_transfer[best_col] = txt

            if txt == 'R$':
                continue
            if re.match(r'^[\d.,]+$', txt) and abs(x - valor_x) < 40:
                current_transfer['valor'] = _parse_currency(txt)

            if x < contrato_x - 5 and not re.match(r'^\d', txt):
                if current_transfer['cliente']:
                    current_transfer['cliente'] += ' ' + txt
                else:
                    current_transfer['cliente'] = txt

    if current_transfer and current_transfer.get('contrato'):
        transfers.append(current_transfer)

    return transfers


def _parse_pdf_fallback(pdf_path):
    transfers = []
    try:
        with open(pdf_path, 'rb') as f:
            content = f.read()
        text_parts = []
        for match in re.finditer(rb'\((.*?)\)', content):
            try:
                text_parts.append(match.group(1).decode('latin-1', errors='ignore'))
            except:
                pass
        full_text = ' '.join(text_parts)
        transfers.extend(_parse_repom_text(full_text))
    except Exception as e:
        print(f"  ERRO no fallback PDF: {e}")
    return transfers


def _parse_repom_text(text):
    transfers = []
    lines = text.split('\n')
    for line in lines:
        contrato_match = re.search(r'(\d{7,})', line)
        valor_match = re.search(r'R\$\s*([\d.,]+)', line)
        date_matches = re.findall(r'(\d{2}/\d{2}/\d{4})', line)
        if contrato_match and valor_match:
            transfer = {
                'contrato': contrato_match.group(1),
                'valor': _parse_currency(valor_match.group(1)),
                'cliente': '',
                'data_pagamento': '',
                'data_transferencia': '',
            }
            motz_match = re.search(r'(MOTZ\s+TRANSPORTES\s*\w*)', line, re.IGNORECASE)
            if motz_match:
                transfer['cliente'] = motz_match.group(1)
            if len(date_matches) >= 2:
                transfer['data_transferencia'] = date_matches[-1]
                if len(date_matches) >= 3:
                    transfer['data_pagamento'] = date_matches[-2]
            transfers.append(transfer)
    return transfers


def _parse_currency(value_str):
    if isinstance(value_str, (int, float)):
        return float(value_str)
    s = str(value_str).strip().replace('R$', '').strip()
    s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0def read_motz_xlsx(filepath):
    print(f"  Lendo MOTZ: {filepath}")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    records = []
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column - 1

    print(f"  Colunas encontradas: {list(headers.keys())}")

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue

        def get_val(col_name, default=None):
            idx = headers.get(col_name)
            if idx is not None and idx < len(row):
                return row[idx]
            return default

        nf_cliente = get_val('NF cliente')
        if nf_cliente is None:
            continue

        cte_raw = str(get_val('Nº do CTe', '') or '')
        ctes = list(set(c.strip() for c in cte_raw.split(',') if c.strip()))

        record = {
            'nf_cliente': str(nf_cliente).strip(),
            'cte_numeros': ctes,
            'cte_raw': cte_raw,
            'total_cte_bruto': _safe_float(get_val('Total CTe bruto', 0)),
            'data_emissao': str(get_val('Data emissão', '') or ''),
            'carta_frete': str(get_val('Nº carta frete', '') or ''),
            'formulario': str(get_val('Nº formulário', '') or ''),
            'serie': str(get_val('Série', '') or ''),
            'romaneio': str(get_val('Nº romaneio', '') or ''),
            'motorista': str(get_val('Nome motorista', '') or ''),
            'cpf_motorista': str(get_val('C.N.P.J./C.P.F. motorista', '') or ''),
            'cavalo': str(get_val('Cavalo', '') or ''),
            'carreta': str(get_val('Carreta', '') or ''),
            'proprietario': str(get_val('Proprietário', '') or ''),
            'recebedor_frete': str(get_val('Recebedor do Frete', '') or ''),
            'vlr_frete_liquido': _safe_float(get_val('Vlr. frete líquido', 0)),
            'vlr_adiantamento': _safe_float(get_val('Vlr. adiantamento', 0)),
            'vlr_saldo': _safe_float(get_val('Vlr. saldo', 0)),
            'situacao_adto': str(get_val('Situação adto.', '') or ''),
            'situacao_saldo': str(get_val('Situação saldo', '') or ''),
            'data_quitacao': str(get_val('Data quitação', '') or ''),
            'centro_custo': str(get_val('Centro custo', '') or ''),
            'vlr_mercadoria': _safe_float(get_val('Vlr. da mercadoria', 0)),
        }
        records.append(record)

    wb.close()
    return records


def _safe_float(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    try:
        s = str(val).replace(',', '.').strip()
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def read_atua_xls(filepath):
    print(f"  Lendo ATUA: {filepath}")
    ext = Path(filepath).suffix.lower()

    if ext == '.xlsx':
        return _read_atua_xlsx(filepath)

    try:
        import subprocess, shutil
        temp_dir = '/tmp/atua_convert'
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
        os.makedirs(temp_dir, exist_ok=True)
        env = os.environ.copy()
        env['HOME'] = '/tmp'
        result = subprocess.run(
            ['libreoffice', '--headless', '--convert-to', 'xlsx', '--outdir', temp_dir, filepath],
            capture_output=True, text=True, timeout=60, env=env
        )
        xlsx_path = os.path.join(temp_dir, Path(filepath).stem + '.xlsx')
        if os.path.exists(xlsx_path):
            headers, records = _read_atua_xlsx(xlsx_path)
            print(f"  Colunas ATUA (LibreOffice): {headers}")
            return headers, records
        else:
            print(f"  LibreOffice nao gerou arquivo")
    except Exception as e:
        print(f"  LibreOffice conversion failed: {e}")

    try:
        import xlrd
        wb = xlrd.open_workbook(filepath, ignore_workbook_corruption=True)
        ws = wb.sheet_by_index(0)
        headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
        records = []
        for r in range(1, ws.nrows):
            row_dict = {}
            for c in range(ws.ncols):
                row_dict[headers[c]] = ws.cell_value(r, c)
            records.append(row_dict)
        print(f"  Colunas ATUA (xlrd): {headers}")
        return headers, records
    except Exception as e:
        print(f"  xlrd failed: {e}")

    print("  Usando parser BIFF8 embutido...")
    from parse_xls import read_xls
    headers, rows = read_xls(filepath)
    print(f"  Colunas ATUA (BIFF): {headers}")
    return headers, rows


def _read_atua_xlsx(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    expected_cols = {'nr_titulo', 'nr_nf', 'nr_ctrc', 'vl_total', 'vl_frete'}
    row1 = [str(c.value or '').strip() for c in ws[1]]
    data_start = 2
    if not expected_cols.intersection(set(h.lower() for h in row1)):
        row2 = [str(c.value or '').strip() for c in ws[2]]
        if expected_cols.intersection(set(h.lower() for h in row2)):
            print(f"  Primeira linha ignorada (totais), usando linha 2 como cabecalho")
            row1 = row2
            data_start = 3
    headers = row1
    records = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not row or all(v is None for v in row):
            continue
        row_dict = {}
        for i, h in enumerate(headers):
            if i < len(row):
                row_dict[h] = row[i]
        records.append(row_dict)
    wb.close()
    print(f"  Colunas ATUA (xlsx): {headers}")
    return headers, records


def reconcile(motz_records, atua_headers, atua_records, pdf_transfers, quebra_records=None):
    """v4.1 FIX: quebra NFs separadas por virgula no nr_nf ATUA para match correto."""

    quebra_by_nf = {}
    quebra_source = quebra_records if quebra_records else atua_records
    for rec in quebra_source:
        nf_raw = rec.get('nr_nf', '')
        vq = _safe_float(rec.get('vl_quebra_avaria', 0))
        if vq == 0:
            continue
        nfs = _split_nf_list(nf_raw)
        for nf_clean in nfs:
            if nf_clean not in quebra_by_nf:
                quebra_by_nf[nf_clean] = vq
            else:
                quebra_by_nf[nf_clean] += vq

    atua_by_titulo = {}
    atua_by_ctrc = {}
    atua_by_nf = {}

    for rec in atua_records:
        titulo_raw = str(rec.get('nr_titulo', '') or '').strip()
        ctrc_raw = str(rec.get('nr_ctrc', '') or '').strip()
        nf_raw = rec.get('nr_nf', '')
        vl_frete = _safe_float(rec.get('vl_total', 0))

        titulo_clean = titulo_raw.split('.')[0] if '.' in titulo_raw else titulo_raw
        ctrc_clean = ctrc_raw.split('.')[0] if '.' in ctrc_raw else ctrc_raw

        nfs_clean = _split_nf_list(nf_raw)
        nf_display = ', '.join(nfs_clean) if nfs_clean else ''

        if titulo_clean:
            atua_by_titulo[titulo_clean] = {
                'nr_titulo': titulo_clean,
                'nr_ctrc': ctrc_clean,
                'vl_frete': vl_frete,
                'nr_nf': nf_display,
                'nm_pessoa_matriz': str(rec.get('nm_pessoa_matriz', '') or ''),
                'nr_cpf_cnpj_raiz': str(rec.get('nr_cpf_cnpj_raiz', '') or ''),
            }

        if ctrc_clean:
            atua_by_ctrc[ctrc_clean] = atua_by_titulo.get(titulo_clean, {
                'nr_titulo': titulo_clean,
                'nr_ctrc': ctrc_clean,
                'vl_frete': vl_frete,
                'nr_nf': nf_display,
                'nm_pessoa_matriz': str(rec.get('nm_pessoa_matriz', '') or ''),
                'nr_cpf_cnpj_raiz': str(rec.get('nr_cpf_cnpj_raiz', '') or ''),
            })

        for nf_clean in nfs_clean:
            atua_entry = {
                'nr_titulo': titulo_clean,
                'nr_ctrc': ctrc_clean,
                'vl_frete': vl_frete,
                'nr_nf': nf_clean,
                'nm_pessoa_matriz': str(rec.get('nm_pessoa_matriz', '') or ''),
                'nr_cpf_cnpj_raiz': str(rec.get('nr_cpf_cnpj_raiz', '') or ''),
            }
            if nf_clean not in atua_by_nf:
                atua_by_nf[nf_clean] = atua_entry
            else:
                existing = atua_by_nf[nf_clean]
                if isinstance(existing, dict):
                    existing['vl_frete'] = existing['vl_frete'] + vl_frete
                    if ctrc_clean and ctrc_clean not in existing.get('nr_ctrc', ''):
                        existing['nr_ctrc'] = existing['nr_ctrc'] + ',' + ctrc_clean

    pdf_by_contrato = {}
    for t in pdf_transfers:
        contrato = str(t.get('contrato', '')).strip()
        if contrato:
            pdf_by_contrato.setdefault(contrato, []).append(t)

    print(f"\n  Indices criados:")
    print(f"    ATUA por titulo: {len(atua_by_titulo)} registros")
    print(f"    ATUA por CTRC: {len(atua_by_ctrc)} registros")
    print(f"    ATUA por NF: {len(atua_by_nf)} registros (inclui NFs separadas por virgula)")
    print(f"    PDF por contrato: {len(pdf_by_contrato)} registros")

    results = []
    matched_atua = set()
    matched_pdf = set()

    for motz in motz_records:
        nf = motz['nf_cliente']
        formulario = motz['formulario']
        ctes = motz['cte_numeros']

        soma_adto_saldo = round(motz['vlr_adiantamento'] + motz['vlr_saldo'], 2)
        frete_liq = round(motz['vlr_frete_liquido'], 2)
        divergencia_interna = round(frete_liq - soma_adto_saldo, 2)

        atua_match = None
        nf_clean_list = _split_nf_list(nf)

        for nf_val in nf_clean_list:
            if nf_val in atua_by_nf:
                atua_match = atua_by_nf[nf_val]
                matched_atua.add(nf_val)
                break
            if nf_val in atua_by_titulo:
                atua_match = atua_by_titulo[nf_val]
                matched_atua.add(nf_val)
                break

        if atua_match is None:
            for cte in ctes:
                cte_clean = cte.split('.')[0] if '.' in cte else cte
                cte_clean = cte_clean.lstrip('0') or '0'
                if cte_clean in atua_by_ctrc:
                    atua_match = atua_by_ctrc[cte_clean]
                    matched_atua.add(cte_clean)
                    break

        vl_quebra_avaria = 0
        for nf_val in nf_clean_list:
            if nf_val in quebra_by_nf:
                vl_quebra_avaria = quebra_by_nf[nf_val]
                break
        divergencia_interna = round(divergencia_interna - vl_quebra_avaria, 2)

        vl_frete_atua = atua_match['vl_frete'] if atua_match else None
        diferenca = None
        status = 'NÃO ENCONTRADO'

        if vl_frete_atua is not None:
            diferenca = round(vl_frete_atua - frete_liq, 2)
            if abs(diferenca) <= 0.02:
                status = 'OK'
                diferenca = 0.0
            elif diferenca < 0:
                status = 'ATUA MENOR'
            else:
                status = 'ATUA MAIOR'

        ctrc_final = ''
        for nf_val in nf_clean_list:
            if nf_val in atua_by_nf:
                ctrc_val = atua_by_nf[nf_val].get('nr_ctrc', '')
                if ctrc_val:
                    ctrc_final = ctrc_val
                    break
        if not ctrc_final and atua_match and atua_match.get('nr_ctrc'):
            ctrc_final = atua_match['nr_ctrc']
        if not ctrc_final and ctes:
            ctrc_final = ', '.join(ctes)

        pdf_matches = pdf_by_contrato.get(formulario, [])
        if pdf_matches:
            matched_pdf.add(formulario)

        nr_ctrc_atua = ''
        for nf_val in nf_clean_list:
            if nf_val in atua_by_nf:
                nr_ctrc_atua = atua_by_nf[nf_val].get('nr_ctrc', '')
                break

        base = {
            'cliente': motz.get('centro_custo', ''),
            'contrato': formulario,
            'titulo_nfe': nf,
            'ctrc': ctrc_final,
            'nr_ctrc_atua': nr_ctrc_atua,
            'carta_frete': motz['carta_frete'],
            'motorista': motz['motorista'],
            'romaneio': motz.get('romaneio', ''),
            'data_emissao_motz': motz.get('data_emissao', ''),
            'vlr_frete_liquido': frete_liq,
            'vlr_adiantamento': motz['vlr_adiantamento'],
            'vlr_saldo': motz['vlr_saldo'],
            'soma_adto_saldo': soma_adto_saldo,
            'vl_quebra_avaria': vl_quebra_avaria,
            'divergencia_interna': divergencia_interna if abs(divergencia_interna) > 0.02 else 0.0,
            'vl_frete_atua': vl_frete_atua,
            'diferenca': diferenca,
            'status': status,
            'situacao_adto': motz['situacao_adto'],
            'situacao_saldo': motz['situacao_saldo'],
        }

        if pdf_matches:
            for pdf_m in pdf_matches:
                row = dict(base)
                row['has_pdf'] = True
                row['data_emissao_repom'] = pdf_m.get('data_emissao', '')
                row['data_quitacao_repom'] = pdf_m.get('data_quitacao', '')
                row['data_pagamento'] = pdf_m.get('data_pagamento', '')
                row['data_transferencia'] = pdf_m.get('data_transferencia', '')
                row['valor_transferido'] = pdf_m.get('valor', '')
                results.append(row)
        else:
            row = dict(base)
            row['has_pdf'] = False
            row['data_emissao_repom'] = ''
            row['data_quitacao_repom'] = ''
            row['data_pagamento'] = ''
            row['data_transferencia'] = ''
            row['valor_transferido'] = ''
            results.append(row)

    seen = set()
    unique_results = []
    duplicates_removed = 0
    for row in results:
        key = tuple(sorted(row.items()))
        if key not in seen:
            seen.add(key)
            unique_results.append(row)
        else:
            duplicates_removed += 1
    if duplicates_removed > 0:
        print(f"  Linhas duplicadas removidas: {duplicates_removed}")
    results = unique_results

    def parse_date_for_sort(rec):
        d = rec.get('data_emissao_motz', '') or ''
        if isinstance(d, datetime):
            return d
        d = str(d).strip()
        for fmt in ('%d/%m/%Y %H:%M:%S', '%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y'):
            try:
                return datetime.strptime(d, fmt)
            except (ValueError, TypeError):
                continue
        return datetime(9999, 12, 31)
    results.sort(key=parse_date_for_sort, reverse=True)

    unmatched_atua = []
    for rec in atua_records:
        titulo = str(rec.get('nr_titulo', '') or '').strip()
        titulo_clean = titulo.split('.')[0] if '.' in titulo else titulo
        ctrc = str(rec.get('nr_ctrc', '') or '').strip()
        ctrc_clean = ctrc.split('.')[0] if '.' in ctrc else ctrc
        nfs_clean = _split_nf_list(rec.get('nr_nf', ''))
        if (titulo_clean not in matched_atua and
            ctrc_clean not in matched_atua and
            not any(n in matched_atua for n in nfs_clean)):
            unmatched_atua.append(rec)

    unmatched_pdf = []
    for t in pdf_transfers:
        contrato = str(t.get('contrato', '')).strip()
        if contrato not in matched_pdf:
            unmatched_pdf.append(t)

    stats = {
        'total_motz': len(motz_records),
        'total_atua': len(atua_records),
        'total_pdf': len(pdf_transfers),
        'matched_atua': len([r for r in results if r['status'] != 'NÃO ENCONTRADO']),
        'total_linhas': len(results),
        'matched_pdf': len(matched_pdf),
        'divergencias': len(set(r['contrato'] for r in results if r['status'] in ('ATUA MENOR', 'ATUA MAIOR'))),
        'ok': len(set(r['contrato'] for r in results if r['status'] == 'OK')),
        'nao_encontrado_atua': len(set(r['contrato'] for r in results if r['status'] == 'NÃO ENCONTRADO')),
        'unmatched_atua': len(unmatched_atua),
        'unmatched_pdf': len(unmatched_pdf),
        'soma_frete_motz': sum(r['vlr_frete_liquido'] for r in results),
        'soma_frete_atua': sum(r['vl_frete_atua'] or 0 for r in results),
    }

    return results, unmatched_atua, unmatched_pdf, statsdef generate_excel(results, unmatched_atua, unmatched_pdf, stats, output_path):
    print(f"\n  Gerando Excel: {output_path}")
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Conciliação Completa"

    columns = [
        ('Cliente', 30), ('Contrato', 14), ('TITULO (NFe)', 14),
        ('CTRC', 16), ('nr_ctrc ATUA', 14), ('Nº Carta Frete', 14),
        ('Motorista', 30), ('Nº Romaneio', 14), ('Data Emissão', 16),
        ('Vlr. Frete Líquido', 18), ('Vlr. Adiantamento', 18), ('Vlr. Saldo', 15),
        ('Soma Adto+Saldo', 18), ('vl_quebra_avaria', 18), ('Diverg. Interna', 16),
        ('vl_total ATUA', 16), ('Diferença MOTZ×ATUA', 20), ('Status', 18),
        ('Data Emissão Repom', 18), ('Data Transferência', 20), ('Valor Transferido', 16),
        ('Situação Adto', 14), ('Situação Saldo', 14),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, 1):
        cell = ws1.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = col_width

    ws1.row_dimensions[1].height = 30
    ws1.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(columns))}1"

    for row_idx, rec in enumerate(results, 2):
        values = [
            rec['cliente'], rec['contrato'], rec['titulo_nfe'],
            rec['ctrc'], rec['nr_ctrc_atua'], rec['carta_frete'],
            rec['motorista'], rec.get('romaneio', ''), rec['data_emissao_motz'],
            rec['vlr_frete_liquido'], rec['vlr_adiantamento'], rec['vlr_saldo'],
            rec['soma_adto_saldo'], rec['vl_quebra_avaria'], rec['divergencia_interna'],
            rec['vl_frete_atua'] if rec['vl_frete_atua'] is not None else '',
            rec['diferenca'] if rec['diferenca'] is not None else '',
            rec['status'], rec['data_emissao_repom'], rec['data_transferencia'],
            rec['valor_transferido'], rec['situacao_adto'], rec['situacao_saldo'],
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            if col_idx in (10, 11, 12, 13, 14, 15, 16, 17, 21):
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0.00'

        status = rec['status']
        diff_cell = ws1.cell(row=row_idx, column=17)
        status_cell = ws1.cell(row=row_idx, column=18)
        atua_cell = ws1.cell(row=row_idx, column=16)

        if status == 'OK':
            for c in (diff_cell, status_cell, atua_cell):
                c.fill = GREEN_FILL
                c.font = GREEN_FONT
        elif status in ('ATUA MENOR', 'ATUA MAIOR'):
            diff_val = abs(rec['diferenca']) if rec['diferenca'] is not None else 0
            if diff_val > 100:
                for c in (diff_cell, status_cell, atua_cell):
                    c.fill = RED_FILL
                    c.font = RED_FONT
            elif status == 'ATUA MENOR':
                for c in (diff_cell, status_cell, atua_cell):
                    c.fill = RED_FILL
                    c.font = RED_FONT
            else:
                for c in (diff_cell, status_cell, atua_cell):
                    c.fill = BLUE_FILL
                    c.font = BLUE_FONT
        elif status == 'NÃO ENCONTRADO':
            for c in (diff_cell, status_cell):
                c.fill = YELLOW_FILL

        div_interna = abs(rec['divergencia_interna']) if rec['divergencia_interna'] else 0
        if div_interna > 100:
            ws1.cell(row=row_idx, column=15).fill = RED_FILL
            ws1.cell(row=row_idx, column=15).font = RED_FONT
        elif rec['divergencia_interna'] != 0.0:
            ws1.cell(row=row_idx, column=15).fill = YELLOW_FILL
            ws1.cell(row=row_idx, column=15).font = Font(color="9C6500", bold=True)

        if not rec.get('has_pdf', True):
            for col_idx in range(1, len(columns) + 1):
                cell = ws1.cell(row=row_idx, column=col_idx)
                if cell.fill == PatternFill() or cell.fill is None:
                    cell.fill = YELLOW_FILL

        sit_saldo = str(rec.get('situacao_saldo', '') or '').strip()
        if sit_saldo.upper() == 'ABERTO':
            for col_idx in (12, 23):
                cell = ws1.cell(row=row_idx, column=col_idx)
                cell.fill = PURPLE_FILL
                cell.font = PURPLE_FONT

    ws2 = wb.create_sheet("Resumo")
    ws2.column_dimensions['A'].width = 40
    ws2.column_dimensions['B'].width = 20

    summary_data = [
        ("CONCILIAÇÃO BANCÁRIA MOTZ TRANSPORTES", ""),
        (f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ""),
        ("", ""),
        ("TOTAIS POR FONTE", "Qtd"),
        ("Registros no MOTZ", stats['total_motz']),
        ("Registros no ATUA", stats['total_atua']),
        ("Transferências no PDF Repom", stats['total_pdf']),
        ("", ""),
        ("RESULTADO DA CONCILIAÇÃO", "Qtd"),
        ("Conciliados OK (valores batem)", stats['ok']),
        ("Com divergência MOTZ × ATUA", stats['divergencias']),
        ("Sem correspondência no ATUA", stats['nao_encontrado_atua']),
        ("Registros ATUA sem correspondência no MOTZ", stats['unmatched_atua']),
        ("Transferências PDF sem correspondência", stats['unmatched_pdf']),
        ("", ""),
        ("VALORES", "R$"),
        ("Soma Frete Líquido (MOTZ)", stats['soma_frete_motz']),
        ("Soma vl_total (ATUA)", stats['soma_frete_atua']),
        ("Diferença Total", stats['soma_frete_atua'] - stats['soma_frete_motz']),
    ]

    for row_idx, (label, value) in enumerate(summary_data, 1):
        cell_a = ws2.cell(row=row_idx, column=1, value=label)
        cell_b = ws2.cell(row=row_idx, column=2, value=value)
        if row_idx in (1,):
            cell_a.font = Font(bold=True, size=14)
        elif label in ("TOTAIS POR FONTE", "RESULTADO DA CONCILIAÇÃO", "VALORES"):
            cell_a.font = Font(bold=True, size=11)
            cell_a.fill = HEADER_FILL
            cell_a.font = HEADER_FONT
            cell_b.fill = HEADER_FILL
            cell_b.font = HEADER_FONT
        if isinstance(value, float):
            cell_b.number_format = '#,##0.00'

    ws3 = wb.create_sheet("Não Encontrados")
    ws3.cell(row=1, column=1, value="REGISTROS MOTZ SEM CORRESPONDÊNCIA NO ATUA").font = Font(bold=True, size=12)
    ws3.merge_cells('A1:F1')

    nao_encontrados_motz = [r for r in results if r['status'] == 'NÃO ENCONTRADO']
    if nao_encontrados_motz:
        headers_ne = ['TITULO (NFe)', 'Contrato', 'CTRC', 'Vlr. Frete Líquido', 'Motorista', 'Cliente']
        for col_idx, h in enumerate(headers_ne, 1):
            cell = ws3.cell(row=2, column=col_idx, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
        for row_idx, rec in enumerate(nao_encontrados_motz, 3):
            vals = [rec['titulo_nfe'], rec['contrato'], rec['ctrc'],
                    rec['vlr_frete_liquido'], rec['motorista'], rec['cliente']]
            for col_idx, v in enumerate(vals, 1):
                cell = ws3.cell(row=row_idx, column=col_idx, value=v)
                cell.border = THIN_BORDER
                if isinstance(v, float):
                    cell.number_format = '#,##0.00'

    start_row = len(nao_encontrados_motz) + 5
    ws3.cell(row=start_row, column=1, value="REGISTROS ATUA SEM CORRESPONDÊNCIA NO MOTZ").font = Font(bold=True, size=12)
    ws3.merge_cells(f'A{start_row}:F{start_row}')

    if unmatched_atua:
        atua_ne_headers = ['nr_titulo', 'nr_ctrc', 'vl_total', 'nr_nf', 'nm_pessoa_matriz']
        for col_idx, h in enumerate(atua_ne_headers, 1):
            cell = ws3.cell(row=start_row + 1, column=col_idx, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
        for row_idx, rec in enumerate(unmatched_atua, start_row + 2):
            vals = [rec.get('nr_titulo', ''), rec.get('nr_ctrc', ''),
                    _safe_float(rec.get('vl_total', 0)), rec.get('nr_nf', ''),
                    rec.get('nm_pessoa_matriz', '')]
            for col_idx, v in enumerate(vals, 1):
                cell = ws3.cell(row=row_idx, column=col_idx, value=v)
                cell.border = THIN_BORDER
                if isinstance(v, float):
                    cell.number_format = '#,##0.00'

    for col_letter, width in [('A', 18), ('B', 16), ('C', 16), ('D', 18), ('E', 30), ('F', 30)]:
        ws3.column_dimensions[col_letter].width = width

    wb.save(output_path)
    print(f"  Arquivo salvo: {output_path}")
    return output_path


def main():
    parser = argparse.ArgumentParser(description='Conciliação Bancária MOTZ TRANSPORTES')
    parser.add_argument('--motz', required=True)
    parser.add_argument('--atua', required=True)
    parser.add_argument('--atua-quebra', default=None)
    parser.add_argument('--pdfs', nargs='+', default=[])
    parser.add_argument('--output', default='conciliacao_motz.xlsx')
    args = parser.parse_args()

    print("=" * 60)
    print("  CONCILIAÇÃO BANCÁRIA — MOTZ TRANSPORTES  (v4.1)")
    print("=" * 60)

    print("\n[1/4] Lendo fontes de dados...")
    motz_records = read_motz_xlsx(args.motz)
    print(f"  -> {len(motz_records)} registros MOTZ")

    atua_headers, atua_records = read_atua_xls(args.atua)
    print(f"  -> {len(atua_records)} registros ATUA")

    quebra_records = []
    if args.atua_quebra:
        print(f"  Lendo ATUA quebra: {args.atua_quebra}")
        _, quebra_records = read_atua_xls(args.atua_quebra)
        print(f"  -> {len(quebra_records)} registros ATUA quebra")

    pdf_transfers = []
    if args.pdfs:
        pdf_transfers = parse_pdf_repom(args.pdfs)
    print(f"  -> {len(pdf_transfers)} transferencias PDF")

    print("\n[2/4] Executando conciliacao...")
    results, unmatched_atua, unmatched_pdf, stats = reconcile(
        motz_records, atua_headers, atua_records, pdf_transfers, quebra_records
    )

    print("\n[3/4] Gerando planilha Excel...")
    output_path = generate_excel(results, unmatched_atua, unmatched_pdf, stats, args.output)

    print("\n[4/4] RESUMO")
    print("-" * 40)
    print(f"  Registros MOTZ:       {stats['total_motz']}")
    print(f"  Registros ATUA:       {stats['total_atua']}")
    print(f"  Conciliados OK:       {stats['ok']}")
    print(f"  Divergencias:         {stats['divergencias']}")
    print(f"  Sem match no ATUA:    {stats['nao_encontrado_atua']}")
    print(f"  Soma MOTZ:            R$ {stats['soma_frete_motz']:,.2f}")
    print(f"  Soma ATUA:            R$ {stats['soma_frete_atua']:,.2f}")
    print("-" * 40)
    print(f"  Arquivo gerado: {output_path}")
    print("=" * 60)

    return output_path


if __name__ == '__main__':
    main()
