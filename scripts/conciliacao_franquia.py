#!/usr/bin/env python3
"""
Conciliação Franquia × Franqueadora — ADAPTA × PIANETTO

Cruza:
- ADAPTA (franqueadora): XLS de Cobrança ATUA com CTRCs faturados pra Pianetto
- PIANETTO (franquia):   PDF "Título Débito × Contrato de Frete" com pagamentos
                         efetuados pela Pianetto à Adapta

Chave de cruzamento: nr_doc_anterior (Adapta, split por "/") = Nr. Doc. (PDF Pianetto)

Gera:
- Excel visual com cores (igual ao MOTZ)
- CSV planilha_padrao.csv pra importação no ATUA Pianetto

v1.0
"""
import io
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import pdfplumber

# ============================================================
# Utilitários
# ============================================================
def parse_currency(s):
    """Converte '4.206,25' em 4206.25 (formato BR)."""
    if s is None or s == "" or s == "0" or s == "0,00":
        return 0.0
    if isinstance(s, (int, float)):
        return float(s)
    s = str(s).replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def fmt_atua(v):
    """Formata número pro padrão aceito pelo ATUA na importação."""
    if v is None or v == "" or v == 0:
        return "0"
    try:
        f = float(v)
    except (ValueError, TypeError):
        return "0"
    if f == 0:
        return "0"
    if f == int(f):
        return str(int(f))
    return f"{f:.2f}".replace(".", ",")


def fmt_rs(v):
    """Formata como R$ X.XXX,YY (pra exibição)."""
    if v is None:
        return "—"
    try:
        f = float(v)
    except (ValueError, TypeError):
        return "—"
    sig = "-" if f < 0 else ""
    abs_n = abs(f)
    return f"{sig}R$ {abs_n:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


# ============================================================
# Parser PDF Pianetto (Título Débito × Contrato de Frete)
# ============================================================
# Linhas de adiantamento NÃO têm peso saída/chegada
RE_ADTO = re.compile(
    r'^ADAPTA\s+\S+\s+(.+?)\s+'                # filial, data_rec, usuario+motorista (lazy)
    r'(\d{2}/\d{2}/\d{2,4})\s+'                # data emissão CF
    r'([A-Z]{3}\d[A-Z0-9]\d{2}|[A-Z]{3}\d{4})\s+'  # placa (Mercosul ou antiga)
    r'(\d+)\s+'                                 # Nr. Doc
    r'(\d+)\s+'                                 # Parcela
    r'Adiantam\s+'                              # Tipo
    r'([\d.,]+)\s+'                             # Valor (bruto)
    r'([\d.,]+)\s+'                             # Pedágio
    r'([\d.,]+)\s+'                             # Quebra
    r'([\d.,]+)\s+'                             # Outros
    r'([\d.,]+)'                                # Total (líquido)
)
# Linhas de saldo TÊM peso saída + chegada
RE_SALDO = re.compile(
    r'^ADAPTA\s+\S+\s+(.+?)\s+'
    r'(\d{2}/\d{2}/\d{2,4})\s+'
    r'([\d.,]+)\s+'                             # Peso saída
    r'([\d.,]+)\s+'                             # Peso chegada
    r'([A-Z]{3}\d[A-Z0-9]\d{2}|[A-Z]{3}\d{4})\s+'
    r'(\d+)\s+'
    r'(\d+)\s+'
    r'Saldo\s+'
    r'([\d.,]+)\s+'
    r'([\d.,]+)\s+'
    r'([\d.,]+)\s+'
    r'([\d.,]+)\s+'
    r'([\d.,]+)'
)
RE_TITULO_MOV = re.compile(r'Título\s+(\d+)\s+Nr')
RE_MOVIMENTO = re.compile(r'Nr\.\s*Movimento\s+(\d+)')
RE_VL_MOV = re.compile(r'Vl\.\s*Movimento\s+([\d.,]+)')
RE_DT_MOV = re.compile(r'Dt\.\s*Movimento\s+(\d{2}/\d{2}/\d{4})')


def read_pianetto_pdf(pdf_bytes):
    """
    Lê o PDF de Título Débito × Contrato de Frete da Pianetto.
    Retorna dict com:
      - titulo_mov: número do título financeiro do movimento (ex: 9642)
      - nr_movimento: nº do movimento (ex: 4246)
      - vl_movimento: valor total do movimento
      - dt_movimento: data do movimento
      - linhas: lista de dicts com cada carta frete (adto ou saldo)
    """
    if hasattr(pdf_bytes, "read"):
        pdf_source = pdf_bytes
    else:
        pdf_source = io.BytesIO(pdf_bytes)

    result = {
        "titulo_mov": None,
        "nr_movimento": None,
        "vl_movimento": 0.0,
        "dt_movimento": None,
        "linhas": [],
    }

    with pdfplumber.open(pdf_source) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            for line in text.split("\n"):
                line = line.strip()

                # Captura metadados do movimento (uma única vez)
                if result["titulo_mov"] is None:
                    m = RE_TITULO_MOV.search(line)
                    if m:
                        result["titulo_mov"] = m.group(1)
                if result["nr_movimento"] is None:
                    m = RE_MOVIMENTO.search(line)
                    if m:
                        result["nr_movimento"] = m.group(1)
                if not result["vl_movimento"]:
                    m = RE_VL_MOV.search(line)
                    if m:
                        result["vl_movimento"] = parse_currency(m.group(1))
                if result["dt_movimento"] is None:
                    m = RE_DT_MOV.search(line)
                    if m:
                        result["dt_movimento"] = m.group(1)

                if not line.startswith("ADAPTA"):
                    continue

                m = RE_ADTO.match(line)
                if m:
                    motorista_chunk = m.group(1).strip()
                    motorista = motorista_chunk.split(" ", 2)[-1] if " " in motorista_chunk else motorista_chunk
                    result["linhas"].append({
                        "motorista": motorista,
                        "data_emi_cf": m.group(2),
                        "peso_saida": 0.0,
                        "peso_chegada": 0.0,
                        "placa": m.group(3),
                        "nr_doc": m.group(4),
                        "parcela": m.group(5),
                        "tipo": "A",
                        "valor": parse_currency(m.group(6)),
                        "pedagio": parse_currency(m.group(7)),
                        "quebra": parse_currency(m.group(8)),
                        "outros": parse_currency(m.group(9)),
                        "total": parse_currency(m.group(10)),
                    })
                    continue

                m = RE_SALDO.match(line)
                if m:
                    motorista_chunk = m.group(1).strip()
                    motorista = motorista_chunk.split(" ", 2)[-1] if " " in motorista_chunk else motorista_chunk
                    result["linhas"].append({
                        "motorista": motorista,
                        "data_emi_cf": m.group(2),
                        "peso_saida": parse_currency(m.group(3)),
                        "peso_chegada": parse_currency(m.group(4)),
                        "placa": m.group(5),
                        "nr_doc": m.group(6),
                        "parcela": m.group(7),
                        "tipo": "S",
                        "valor": parse_currency(m.group(8)),
                        "pedagio": parse_currency(m.group(9)),
                        "quebra": parse_currency(m.group(10)),
                        "outros": parse_currency(m.group(11)),
                        "total": parse_currency(m.group(12)),
                    })

    return result


# ============================================================
# Parser XLS Adapta (Cobrança ATUA)
# ============================================================
def read_adapta_xls(xls_bytes):
    """
    Lê o XLS de Cobrança ATUA da Adapta.
    Retorna dict indexado por nr_doc (sem parcela) com lista de títulos por doc.
    Cada título tem: nr_titulo, nr_ctrc, vl_total, vl_quebra_avaria, ds_placa,
    nm_motorista, nr_doc_anterior, nr_nf, nm_pessoa_matriz.
    """
    if hasattr(xls_bytes, "read"):
        raw = xls_bytes.read()
    else:
        raw = xls_bytes

    # XLS antigo precisa conversão via LibreOffice (mesma estratégia do MOTZ)
    import tempfile, os, subprocess, shutil
    tmpdir = tempfile.mkdtemp(prefix="franquia_xls_")
    in_path = os.path.join(tmpdir, "input.xls")
    with open(in_path, "wb") as f:
        f.write(raw)

    xlsx_path = None
    # Detecta se já é XLSX
    if raw[:4] == b"PK\x03\x04":
        xlsx_path = in_path
    else:
        try:
            env = os.environ.copy()
            env["HOME"] = tmpdir
            subprocess.run(
                ["libreoffice", "--headless", "--convert-to", "xlsx",
                 "--outdir", tmpdir, in_path],
                capture_output=True, text=True, timeout=120, env=env,
            )
            candidate = os.path.join(tmpdir, "input.xlsx")
            if os.path.exists(candidate):
                xlsx_path = candidate
        except Exception:
            pass

    if xlsx_path is None:
        shutil.rmtree(tmpdir, ignore_errors=True)
        raise RuntimeError("Não foi possível ler o XLS da Adapta (LibreOffice falhou).")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # Detecta cabeçalho: linha 1 = totais; linha 2 = cabeçalho de colunas
    expected = {"nr_titulo", "nr_ctrc", "nr_doc_anterior", "vl_total"}
    row1 = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    data_start = 2
    if not expected.intersection(set(h.lower() for h in row1)):
        row2 = [str(ws.cell(2, c).value or "").strip() for c in range(1, ws.max_column + 1)]
        if expected.intersection(set(h.lower() for h in row2)):
            row1 = row2
            data_start = 3
    headers = row1
    col = {h: i + 1 for i, h in enumerate(headers) if h}

    def get(r, name):
        idx = col.get(name)
        if idx is None:
            return None
        return ws.cell(r, idx).value

    by_doc = {}
    all_records = []
    for r in range(data_start, ws.max_row + 1):
        doc_ant = get(r, "nr_doc_anterior")
        if doc_ant is None or str(doc_ant).strip() == "":
            continue
        # Split por "/" e normaliza
        doc_num = str(doc_ant).split("/")[0].strip()
        if "." in doc_num:
            doc_num = doc_num.split(".")[0]
        if not doc_num.isdigit():
            continue

        rec = {
            "nr_titulo": get(r, "nr_titulo"),
            "nr_ctrc": get(r, "nr_ctrc"),
            "vl_total": float(get(r, "vl_total") or 0),
            "vl_frete": float(get(r, "vl_frete") or 0),
            "vl_quebra_avaria": float(get(r, "vl_quebra_avaria") or 0),
            "nr_nf": get(r, "nr_nf"),
            "ds_placa": get(r, "ds_placa") or "",
            "nm_motorista": get(r, "nm_motorista") or "",
            "nm_pessoa_matriz": get(r, "nm_pessoa_matriz") or "",
            "nr_doc_anterior": str(doc_ant),
            "_doc_num": doc_num,
        }
        by_doc.setdefault(doc_num, []).append(rec)
        all_records.append(rec)

    wb.close()
    shutil.rmtree(tmpdir, ignore_errors=True)

    return {"by_doc": by_doc, "all": all_records}


# ============================================================
# Cruzamento e geração de resultados
# ============================================================
def reconcile_franquia(pianetto_data, adapta_data):
    """
    Cruza Pianetto (PDF) × Adapta (XLS).

    Para cada linha Pianetto, escolhe a entry Adapta com vl_total mais próximo
    do valor bruto da Pianetto (entre as ainda não usadas). Se não houver match
    por nr_doc, marca como NÃO ENCONTRADO.

    Retorna lista de dicts (1 por linha Pianetto) já cruzada.
    """
    by_doc = adapta_data["by_doc"]
    titulo_mov = pianetto_data.get("titulo_mov") or ""

    matched_titulos = set()
    results = []

    for p in pianetto_data["linhas"]:
        doc = p["nr_doc"]
        candidatos = by_doc.get(doc, [])

        pick = None
        if candidatos:
            livres = [c for c in candidatos
                      if (doc, c["nr_titulo"]) not in matched_titulos]
            if not livres:
                livres = candidatos
            # Match por valor bruto mais próximo
            livres.sort(key=lambda c: abs(c["vl_total"] - p["valor"]))
            pick = livres[0]
            matched_titulos.add((doc, pick["nr_titulo"]))

        if pick:
            diff = round(pick["vl_total"] - p["valor"], 2)
            status = "OK" if abs(diff) < 0.01 else (
                "ADAPTA MAIOR" if diff > 0 else "ADAPTA MENOR"
            )
            row = {
                "Nr. Movimento": pianetto_data.get("nr_movimento"),
                "Dt. Movimento": pianetto_data.get("dt_movimento"),
                "Titulo Mov": titulo_mov,
                "Nr. Doc. (Pianetto)": doc,
                "Tipo": p["tipo"],
                "Placa": p["placa"],
                "Motorista": p["motorista"],
                "Data Emissão CF": p["data_emi_cf"],
                "Peso Saída": p["peso_saida"],
                "Peso Chegada": p["peso_chegada"],
                "Valor Bruto Pianetto": p["valor"],
                "Pedágio": p["pedagio"],
                "Quebra Pianetto": p["quebra"],
                "Outros": p["outros"],
                "Total Pago Pianetto": p["total"],
                "nr_titulo Adapta": pick["nr_titulo"],
                "nr_ctrc Adapta": pick["nr_ctrc"],
                "vl_total Adapta": pick["vl_total"],
                "vl_quebra_avaria Adapta": pick["vl_quebra_avaria"],
                "Placa Adapta": pick["ds_placa"],
                "Diferença Adapta×Pianetto": diff,
                "Status": status,
            }
        else:
            row = {
                "Nr. Movimento": pianetto_data.get("nr_movimento"),
                "Dt. Movimento": pianetto_data.get("dt_movimento"),
                "Titulo Mov": titulo_mov,
                "Nr. Doc. (Pianetto)": doc,
                "Tipo": p["tipo"],
                "Placa": p["placa"],
                "Motorista": p["motorista"],
                "Data Emissão CF": p["data_emi_cf"],
                "Peso Saída": p["peso_saida"],
                "Peso Chegada": p["peso_chegada"],
                "Valor Bruto Pianetto": p["valor"],
                "Pedágio": p["pedagio"],
                "Quebra Pianetto": p["quebra"],
                "Outros": p["outros"],
                "Total Pago Pianetto": p["total"],
                "nr_titulo Adapta": None,
                "nr_ctrc Adapta": None,
                "vl_total Adapta": None,
                "vl_quebra_avaria Adapta": None,
                "Placa Adapta": None,
                "Diferença Adapta×Pianetto": None,
                "Status": "NÃO ENCONTRADO",
            }
        results.append(row)

    return results


# ============================================================
# Gerador do CSV padrão (importação no ATUA Pianetto)
# ============================================================
def generate_csv_baixa(results):
    """
    Gera o CSV planilha_padrao.csv exatamente no formato do ATUA.
    Só inclui linhas que deram match (Status != NÃO ENCONTRADO).
    """
    import csv
    buf = io.StringIO()
    fieldnames = [
        "Nr. CTRC", "Serie", "Valor Pago", "Valor Desconto", "Valor de Juros",
        "Valor Desconto Quebra", "Valor Acres. Quebra",
        "Tipo Parcela (A = Adiantamento / S = Saldo)", "Nr. Fatura",
    ]
    writer = csv.DictWriter(buf, fieldnames=fieldnames, delimiter=";",
                            lineterminator="\r\n")
    writer.writeheader()
    n = 0
    for r in results:
        if r["Status"] == "NÃO ENCONTRADO":
            continue
        writer.writerow({
            "Nr. CTRC": str(r["nr_ctrc Adapta"] or "").split(".")[0],
            "Serie": "1",
            "Valor Pago": fmt_atua(r["Total Pago Pianetto"]),
            "Valor Desconto": "0",
            "Valor de Juros": "0",
            "Valor Desconto Quebra": fmt_atua(r["Quebra Pianetto"]),
            "Valor Acres. Quebra": "0",
            "Tipo Parcela (A = Adiantamento / S = Saldo)": r["Tipo"],
            "Nr. Fatura": str(r["nr_titulo Adapta"] or "").split(".")[0],
        })
        n += 1
    return buf.getvalue().encode("utf-8"), n


# ============================================================
# Gerador Excel visual com cores
# ============================================================
COLORS_EXCEL = {
    "OK":             {"bg": "C8E6C9", "fg": "1B5E20"},  # verde
    "ADAPTA MAIOR":   {"bg": "FFCDD2", "fg": "B71C1C"},  # vermelho
    "ADAPTA MENOR":   {"bg": "BBDEFB", "fg": "0D47A1"},  # azul
    "NÃO ENCONTRADO": {"bg": "FFE0B2", "fg": "E65100"},  # laranja
}


def generate_excel_franquia(results, pianetto_data, adapta_data):
    """Gera Excel visual da conciliação."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conciliação Franquia"

    # Cabeçalho do movimento
    ws["A1"] = "Conciliação Franquia × Franqueadora — ADAPTA × PIANETTO"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:I1")

    ws["A2"] = f"Movimento: {pianetto_data.get('nr_movimento','?')}"
    ws["D2"] = f"Data: {pianetto_data.get('dt_movimento','?')}"
    ws["G2"] = f"Valor: {fmt_rs(pianetto_data.get('vl_movimento', 0))}"
    ws["A3"] = f"Título Mov: {pianetto_data.get('titulo_mov','?')}"
    ws["D3"] = f"Linhas Pianetto: {len(results)}"
    sem_match = sum(1 for r in results if r["Status"] == "NÃO ENCONTRADO")
    com_match = len(results) - sem_match
    ws["G3"] = f"Match: {com_match}/{len(results)}"

    # Cabeçalhos da tabela
    headers = [
        "Nr. Movimento", "Dt. Movimento", "Titulo Mov", "Nr. Doc. (Pianetto)",
        "Tipo", "Placa", "Motorista", "Data Emissão CF",
        "Valor Bruto Pianetto", "Quebra Pianetto", "Total Pago Pianetto",
        "nr_titulo Adapta", "nr_ctrc Adapta", "vl_total Adapta",
        "vl_quebra_avaria Adapta", "Diferença Adapta×Pianetto", "Status",
    ]
    row_header = 5
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row_header, column=ci, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8E8E8")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Dados
    money_cols = {"Valor Bruto Pianetto", "Quebra Pianetto", "Total Pago Pianetto",
                  "vl_total Adapta", "vl_quebra_avaria Adapta",
                  "Diferença Adapta×Pianetto"}

    for ri, r in enumerate(results, start=row_header + 1):
        for ci, h in enumerate(headers, 1):
            v = r.get(h)
            cell = ws.cell(row=ri, column=ci, value=v)
            if h in money_cols and v is not None:
                cell.number_format = 'R$ #,##0.00'

        # Colorir linha pelo status
        status = r.get("Status", "")
        color = COLORS_EXCEL.get(status)
        if color:
            fill = PatternFill("solid", fgColor=color["bg"])
            for ci in range(1, len(headers) + 1):
                ws.cell(row=ri, column=ci).fill = fill
                ws.cell(row=ri, column=ci).font = Font(color=color["fg"])

    # Largura das colunas
    widths = {
        "A": 13, "B": 13, "C": 12, "D": 16, "E": 6, "F": 11, "G": 25, "H": 14,
        "I": 17, "J": 14, "K": 17, "L": 14, "M": 14, "N": 14, "O": 18, "P": 18, "Q": 16,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A6"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# Função orquestradora (uso interno do dashboard)
# ============================================================
def rodar_conciliacao_franquia(pdf_bytes, xls_bytes):
    """
    Função principal chamada pelo dashboard. Recebe os bytes dos arquivos
    e retorna o tuple (results, pianetto_data, adapta_data, csv_bytes,
    xlsx_bytes, n_csv_rows).
    """
    pianetto_data = read_pianetto_pdf(pdf_bytes)
    adapta_data = read_adapta_xls(xls_bytes)
    results = reconcile_franquia(pianetto_data, adapta_data)
    csv_bytes, n_csv = generate_csv_baixa(results)
    xlsx_bytes = generate_excel_franquia(results, pianetto_data, adapta_data)
    return results, pianetto_data, adapta_data, csv_bytes, xlsx_bytes, n_csv


# ============================================================
# Execução standalone (teste pela CLI)
# ============================================================
if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--pdf", required=True, help="PDF da Pianetto")
    parser.add_argument("--xls", required=True, help="XLS de Cobrança Adapta")
    parser.add_argument("--out", default="conciliacao_franquia.xlsx")
    parser.add_argument("--csv", default="planilha_padrao.csv")
    args = parser.parse_args()

    with open(args.pdf, "rb") as f:
        pdf_bytes = f.read()
    with open(args.xls, "rb") as f:
        xls_bytes = f.read()

    results, pianetto_data, adapta_data, csv_bytes, xlsx_bytes, n_csv = \
        rodar_conciliacao_franquia(pdf_bytes, xls_bytes)

    with open(args.out, "wb") as f:
        f.write(xlsx_bytes)
    with open(args.csv, "wb") as f:
        f.write(csv_bytes)

    print(f"Linhas Pianetto: {len(results)}")
    print(f"Match: {sum(1 for r in results if r['Status'] != 'NÃO ENCONTRADO')}")
    print(f"Excel: {args.out}")
    print(f"CSV: {args.csv} ({n_csv} linhas)")
